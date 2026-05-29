"""
Urban EO/AI Evaluation Analysis
===============================
Reads the Urban AI evaluation workbook (Experiment Runs sheet), validates data,
produces summary CSVs, and generates publication-ready figures.

Rubric A (0–10): code generation and GEE execution quality (A1–A5).
Rubric B (0–6):  result interpretation quality (B1–B3).

Usage:
    python evaluations/urban/scripts/analyze_model_evaluation.py
    python evaluations/urban/scripts/analyze_model_evaluation.py \\
        --input evaluations/urban/Urban_AI_Evaluation_Workbook.xlsx
"""

from __future__ import annotations

import argparse
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=FutureWarning)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

URBAN_DIR = Path(__file__).resolve().parents[1]
RESULTS_DIR = URBAN_DIR / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_WORKBOOK = URBAN_DIR / "Urban_AI_Evaluation_Workbook.xlsx"
EXPERIMENT_SHEET = "Experiment Runs"

# ---------------------------------------------------------------------------
# Optional seaborn import
# ---------------------------------------------------------------------------

try:
    import seaborn as sns  # type: ignore

    HAS_SEABORN = True
    sns.set_theme(style="whitegrid", palette="muted")
except ImportError:
    HAS_SEABORN = False

import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

matplotlib.rcParams.update(
    {
        "figure.dpi": 150,
        "savefig.dpi": 150,
        "font.size": 10,
        "axes.titlesize": 12,
        "axes.labelsize": 11,
        "xtick.labelsize": 9,
        "ytick.labelsize": 9,
        "legend.fontsize": 9,
    }
)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

RAG_COLORS = {"With_RAG": "#2196F3", "Without_RAG": "#FF7043"}
RAG_LABELS = {"With_RAG": "With RAG (ON)", "Without_RAG": "Without RAG (OFF)"}

RUBRIC_A_MAX = 10
RUBRIC_B_MAX = 6

SUB_RUBRIC_A = [
    "a1_execution_success",
    "a2_output_validity",
    "a3_dataset_band",
    "a4_gee_patterns",
    "a5_robustness",
]
SUB_RUBRIC_B = [
    "b1_phenomenon",
    "b2_spatial_patterns",
    "b3_artifacts",
]

SUB_RUBRIC_A_LABELS = {
    "a1_execution_success": "A1 Execution",
    "a2_output_validity": "A2 Output",
    "a3_dataset_band": "A3 Dataset/Band",
    "a4_gee_patterns": "A4 GEE Patterns",
    "a5_robustness": "A5 Robustness",
}
SUB_RUBRIC_B_LABELS = {
    "b1_phenomenon": "B1 Phenomenon",
    "b2_spatial_patterns": "B2 Spatial",
    "b3_artifacts": "B3 Uncertainty",
}

TIER_ORDER = [
    "Tier A — Fundamental EO Interactions",
    "Tier B — Aggregation / Completeness / Export",
    "Tier C — Thermal Remote Sensing Core",
    "Tier D — Urban Heat / Applied Analysis",
]

TIER_SHORT = {
    "Tier A — Fundamental EO Interactions": "Tier A",
    "Tier B — Aggregation / Completeness / Export": "Tier B",
    "Tier C — Thermal Remote Sensing Core": "Tier C",
    "Tier D — Urban Heat / Applied Analysis": "Tier D",
}

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="analyze_model_evaluation.py",
        description=(
            "Analyze urban EO/AI model evaluations from the Urban AI workbook. "
            "Generates summary CSVs and figures for Rubric A/B scores."
        ),
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python evaluations/urban/scripts/analyze_model_evaluation.py\n"
            "  python evaluations/urban/scripts/analyze_model_evaluation.py \\\n"
            "    --input evaluations/urban/Urban_AI_Evaluation_Workbook.xlsx\n"
        ),
    )
    parser.add_argument(
        "--input",
        "-i",
        dest="input_xlsx",
        type=str,
        default=None,
        help="Path to the input .xlsx workbook (default: first .xlsx in evaluations/urban/).",
    )
    return parser.parse_args(argv)


def find_excel_file(directory: Path) -> Path:
    if DEFAULT_WORKBOOK.exists():
        print(f"[INFO] Using: {DEFAULT_WORKBOOK.name}")
        return DEFAULT_WORKBOOK
    files = sorted(directory.glob("*.xlsx"))
    if not files:
        sys.exit(f"ERROR: No .xlsx file found in {directory}")
    chosen = files[0]
    if len(files) > 1:
        print(f"[INFO] Multiple .xlsx files found; using: {chosen.name}")
    else:
        print(f"[INFO] Using: {chosen.name}")
    return chosen


# ---------------------------------------------------------------------------
# Load workbook
# ---------------------------------------------------------------------------

# Raw header → canonical column name
COLUMN_MAP = {
    "run_id": "run_id",
    "task_id": "task_id",
    "tier": "tier",
    "task_name": "task_name",
    "model": "model",
    "rag_mode": "rag_mode",
    "run_status": "run_status",
    "execution_outcome": "execution_outcome",
    "refinement_used": "refinement_used",
    "a1_execution_success": "a1_execution_success",
    "a2_output_validity": "a2_output_validity",
    "a3_dataset_band_correctness": "a3_dataset_band",
    "a4_gee_patterns": "a4_gee_patterns",
    "a5_robustness": "a5_robustness",
    "rubric_a_total": "rubric_a_total",
    "rubric_a_label": "rubric_a_label",
    "b1_phenomenon_identification": "b1_phenomenon",
    "b2_spatial_patterns": "b2_spatial_patterns",
    "b3_artifacts_uncertainty": "b3_artifacts",
    "rubric_b_total": "rubric_b_total",
    "rubric_b_label": "rubric_b_label",
    "failure_category": "failure_category",
    "evaluator_notes": "evaluator_notes",
}


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("/", "_")
        .replace("(", "")
        .replace(")", "")
        .replace("__", "_")
    )


def rag_mode_to_status(mode: object) -> str | None:
    if mode is None or (isinstance(mode, float) and np.isnan(mode)):
        return None
    text = str(mode).strip().upper()
    if text in {"ON", "YES", "TRUE", "1", "WITH_RAG", "RAG"}:
        return "With_RAG"
    if text in {"OFF", "NO", "FALSE", "0", "WITHOUT_RAG", "NO RAG"}:
        return "Without_RAG"
    return None


def load_experiment_runs(xlsx_path: Path) -> pd.DataFrame:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        sys.exit("ERROR: openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if EXPERIMENT_SHEET not in wb.sheetnames:
        sys.exit(
            f"ERROR: Sheet '{EXPERIMENT_SHEET}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[EXPERIMENT_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        sys.exit(f"ERROR: Sheet '{EXPERIMENT_SHEET}' is empty.")

    raw_header = rows[0]
    header_norm = [_norm_header(h) for h in raw_header]

    rename: dict[int, str] = {}
    for i, normed in enumerate(header_norm):
        canonical = COLUMN_MAP.get(normed)
        if canonical:
            rename[i] = canonical

    df = pd.DataFrame(rows[1:])
    df = df.rename(columns=rename)
    df = df[[c for c in df.columns if isinstance(c, str)]]

    # Drop blank separator rows
    df = df.dropna(subset=["run_id"], how="all")

    # RAG status
    df["rag_status"] = df["rag_mode"].map(rag_mode_to_status)

    # Numeric coercion
    score_cols = SUB_RUBRIC_A + SUB_RUBRIC_B + ["rubric_a_total", "rubric_b_total"]
    for col in score_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Success: A1 == 2 means script executes without modification
    if "a1_execution_success" in df.columns:
        df["success"] = df["a1_execution_success"] == 2
    else:
        df["success"] = df["rubric_a_total"] >= 9

    df["model"] = df["model"].astype(str).str.strip()
    df["task_id"] = df["task_id"].astype(str).str.strip()

    print(f"[OK] Loaded {len(df)} experiment runs from '{EXPERIMENT_SHEET}'")
    return df


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def validate_and_print(df: pd.DataFrame) -> None:
    print("\n" + "=" * 60)
    print("DATA VALIDATION")
    print("=" * 60)

    print(f"\nColumns: {list(df.columns)}\n")
    print(f"Total runs: {len(df)}")
    print(f"Models ({df['model'].nunique()}): {sorted(df['model'].unique())}")
    print(f"Tasks  ({df['task_id'].nunique()}): {sorted(df['task_id'].unique())}")

    print("\nRAG status counts:")
    print(df["rag_status"].value_counts(dropna=False).to_string())

    print("\nRow counts by model and rag_status:")
    counts = df.groupby(["model", "rag_status"], dropna=False).size().unstack(fill_value=0)
    print(counts.to_string())

    print("\nRubric A totals per model and rag_status:")
    for (model, rag), grp in df.groupby(["model", "rag_status"], dropna=False):
        vals = grp["rubric_a_total"].dropna().tolist()
        print(
            f"  {model:25s} | {str(rag):12s} | n={len(vals):2d} | "
            f"mean={grp['rubric_a_total'].mean():5.2f} | values={vals}"
        )

    print("\nRubric B totals per model and rag_status:")
    for (model, rag), grp in df.groupby(["model", "rag_status"], dropna=False):
        vals = grp["rubric_b_total"].dropna().tolist()
        print(
            f"  {model:25s} | {str(rag):12s} | n={len(vals):2d} | "
            f"mean={grp['rubric_b_total'].mean():5.2f} | values={vals}"
        )

    rag_modes = df["rag_status"].dropna().unique().tolist()
    if "Without_RAG" not in rag_modes:
        print(
            "\n[WARN] No Without-RAG (OFF) runs found yet. "
            "RAG comparison figures will be skipped until OFF rows are scored."
        )

    print("=" * 60)


# ---------------------------------------------------------------------------
# Summary CSV
# ---------------------------------------------------------------------------


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    grp = df.groupby(["model", "rag_status"], dropna=False)

    agg = grp.agg(
        mean_rubric_a=("rubric_a_total", "mean"),
        median_rubric_a=("rubric_a_total", "median"),
        std_rubric_a=("rubric_a_total", "std"),
        mean_rubric_b=("rubric_b_total", "mean"),
        median_rubric_b=("rubric_b_total", "median"),
        std_rubric_b=("rubric_b_total", "std"),
        success_rate=("success", "mean"),
        strong_a_rate=("rubric_a_total", lambda s: (s >= 9).mean()),
        n=("rubric_a_total", "count"),
    ).reset_index()

    def pivot_metric(metric: str) -> pd.DataFrame:
        return agg.pivot(index="model", columns="rag_status", values=metric)

    metrics = [
        "mean_rubric_a",
        "median_rubric_a",
        "std_rubric_a",
        "mean_rubric_b",
        "median_rubric_b",
        "std_rubric_b",
        "success_rate",
        "strong_a_rate",
    ]
    pivoted = [pivot_metric(m) for m in metrics]
    summary = pd.concat(pivoted, axis=1)
    summary.columns = [f"{m}_{rag}" for m in metrics for rag in pivoted[0].columns]
    summary = summary.reset_index()

    for metric in ["mean_rubric_a", "mean_rubric_b", "success_rate"]:
        col_rag = f"{metric}_With_RAG"
        col_no = f"{metric}_Without_RAG"
        delta_col = f"{metric.replace('mean_', '').replace('_rate', '')}_delta_RAG_minus_no_RAG"
        if col_rag in summary.columns and col_no in summary.columns:
            summary[delta_col] = summary[col_rag] - summary[col_no]

    csv_path = RESULTS_DIR / "model_summary.csv"
    summary.to_csv(csv_path, index=False, float_format="%.4f")
    print(f"\n[OK] model_summary.csv → {csv_path}")
    return summary


def build_tier_summary(df: pd.DataFrame) -> pd.DataFrame:
    tier_agg = (
        df.groupby(["tier", "model", "rag_status"], dropna=False)
        .agg(
            mean_rubric_a=("rubric_a_total", "mean"),
            mean_rubric_b=("rubric_b_total", "mean"),
            n=("rubric_a_total", "count"),
        )
        .reset_index()
    )
    csv_path = RESULTS_DIR / "tier_summary.csv"
    tier_agg.to_csv(csv_path, index=False, float_format="%.4f")
    print(f"[OK] tier_summary.csv → {csv_path}")
    return tier_agg


def build_task_summary(df: pd.DataFrame) -> pd.DataFrame:
    task_agg = (
        df.groupby(["task_id", "task_name", "tier", "model", "rag_status"], dropna=False)
        .agg(
            rubric_a_total=("rubric_a_total", "first"),
            rubric_b_total=("rubric_b_total", "first"),
            rubric_a_label=("rubric_a_label", "first"),
            rubric_b_label=("rubric_b_label", "first"),
        )
        .reset_index()
    )
    csv_path = RESULTS_DIR / "task_summary.csv"
    task_agg.to_csv(csv_path, index=False)
    print(f"[OK] task_summary.csv → {csv_path}")
    return task_agg


# ---------------------------------------------------------------------------
# Figure helpers
# ---------------------------------------------------------------------------


def save_fig(fig: plt.Figure, name: str) -> None:
    path = RESULTS_DIR / name
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    print(f"[OK] {name} → {path}")


def model_order(summary: pd.DataFrame, metric: str = "mean_rubric_a_With_RAG") -> list[str]:
    if metric in summary.columns:
        return summary.sort_values(metric, ascending=False)["model"].tolist()
    fallback = [c for c in summary.columns if c.startswith("mean_rubric_a")]
    if fallback:
        return summary.sort_values(fallback[0], ascending=False)["model"].tolist()
    return sorted(summary["model"].tolist())


def has_rag_comparison(df: pd.DataFrame) -> bool:
    modes = set(df["rag_status"].dropna().unique())
    return "With_RAG" in modes and "Without_RAG" in modes


def grouped_bar(
    ax: plt.Axes,
    models: list[str],
    rag_values: list[float],
    no_rag_values: list[float] | None,
    ylabel: str,
    title: str,
    ylim: tuple | None = None,
) -> None:
    x = np.arange(len(models))
    if no_rag_values is not None:
        width = 0.35
        ax.bar(
            x - width / 2,
            rag_values,
            width,
            label=RAG_LABELS["With_RAG"],
            color=RAG_COLORS["With_RAG"],
            alpha=0.85,
        )
        ax.bar(
            x + width / 2,
            no_rag_values,
            width,
            label=RAG_LABELS["Without_RAG"],
            color=RAG_COLORS["Without_RAG"],
            alpha=0.85,
        )
    else:
        ax.bar(
            x,
            rag_values,
            0.6,
            label=RAG_LABELS["With_RAG"],
            color=RAG_COLORS["With_RAG"],
            alpha=0.85,
        )
    ax.set_xticks(x)
    ax.set_xticklabels(models, rotation=35, ha="right")
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.legend()
    if ylim:
        ax.set_ylim(*ylim)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=False, nbins=6))


def _summary_values(
    summary: pd.DataFrame,
    order: list[str],
    col_rag: str,
    col_no: str | None = None,
) -> tuple[list[float], list[float] | None]:
    s = summary.set_index("model").reindex(order)
    rag_vals = s.get(col_rag, pd.Series(dtype=float)).fillna(0).tolist()
    if col_no and col_no in s.columns:
        no_vals = s.get(col_no, pd.Series(dtype=float)).fillna(0).tolist()
        return rag_vals, no_vals
    return rag_vals, None


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------


def fig_mean_rubric_a(summary: pd.DataFrame, order: list[str], compare_rag: bool) -> None:
    rag_vals, no_vals = _summary_values(
        summary, order, "mean_rubric_a_With_RAG", "mean_rubric_a_Without_RAG" if compare_rag else None
    )
    fig, ax = plt.subplots(figsize=(11, 5))
    title = "Mean Rubric A Score by Model (Code & Execution Quality)"
    if not compare_rag:
        title += "\n(RAG ON only — OFF runs not yet scored)"
    grouped_bar(ax, order, rag_vals, no_vals, "Mean Rubric A (/10)", title, ylim=(0, 10.5))
    ax.axhline(9, color="grey", linestyle="--", linewidth=0.8, label="Strong (≥9)")
    ax.axhline(5, color="grey", linestyle=":", linewidth=0.8, label="Partial (≥5)")
    ax.legend()
    fig.tight_layout()
    save_fig(fig, "mean_rubric_a_by_model.png")


def fig_mean_rubric_b(summary: pd.DataFrame, order: list[str], compare_rag: bool) -> None:
    rag_vals, no_vals = _summary_values(
        summary, order, "mean_rubric_b_With_RAG", "mean_rubric_b_Without_RAG" if compare_rag else None
    )
    fig, ax = plt.subplots(figsize=(11, 5))
    title = "Mean Rubric B Score by Model (Interpretation Quality)"
    if not compare_rag:
        title += "\n(RAG ON only — OFF runs not yet scored)"
    grouped_bar(ax, order, rag_vals, no_vals, "Mean Rubric B (/6)", title, ylim=(0, 6.5))
    ax.axhline(6, color="grey", linestyle="--", linewidth=0.8, label="Strong (6)")
    ax.axhline(3, color="grey", linestyle=":", linewidth=0.8, label="Adequate (≥3)")
    ax.legend()
    fig.tight_layout()
    save_fig(fig, "mean_rubric_b_by_model.png")


def fig_code_vs_interpretation(df: pd.DataFrame) -> None:
    grp = (
        df.groupby(["model", "rag_status"], dropna=False)
        .agg(mean_a=("rubric_a_total", "mean"), mean_b=("rubric_b_total", "mean"))
        .reset_index()
    )

    fig, ax = plt.subplots(figsize=(10, 8))
    markers = {"With_RAG": "o", "Without_RAG": "^"}

    for rag in grp["rag_status"].dropna().unique():
        sub = grp[grp["rag_status"] == rag]
        ax.scatter(
            sub["mean_a"],
            sub["mean_b"],
            label=RAG_LABELS.get(rag, str(rag)),
            color=RAG_COLORS.get(rag, "#888888"),
            marker=markers.get(rag, "o"),
            s=120,
            alpha=0.9,
            edgecolors="white",
            linewidths=0.8,
        )
        for _, row in sub.iterrows():
            ax.annotate(
                row["model"],
                (row["mean_a"], row["mean_b"]),
                textcoords="offset points",
                xytext=(6, 4),
                fontsize=8,
                color="dimgray",
            )

    ax.set_xlabel("Mean Rubric A — Code & Execution (/10)")
    ax.set_ylabel("Mean Rubric B — Interpretation (/6)")
    ax.set_title("Code Quality vs Interpretation Quality by Model")
    ax.set_xlim(0, 10.5)
    ax.set_ylim(0, 6.5)
    ax.axvline(9, color="grey", linestyle="--", alpha=0.5)
    ax.axhline(6, color="grey", linestyle="--", alpha=0.5)
    ax.legend()
    ax.grid(linestyle="--", alpha=0.35)
    fig.tight_layout()
    save_fig(fig, "code_vs_interpretation_scatter.png")


def fig_rubric_a_distribution(df: pd.DataFrame, order: list[str], compare_rag: bool) -> None:
    fig, ax = plt.subplots(figsize=(13, 6))

    if HAS_SEABORN and compare_rag:
        sns.violinplot(
            data=df,
            x="model",
            y="rubric_a_total",
            hue="rag_status",
            order=order,
            palette=RAG_COLORS,
            ax=ax,
            inner="quartile",
            cut=0,
        )
        handles, labels = ax.get_legend_handles_labels()
        ax.legend(handles, [RAG_LABELS.get(l, l) for l in labels])
    elif HAS_SEABORN:
        sns.violinplot(
            data=df,
            x="model",
            y="rubric_a_total",
            order=order,
            color=RAG_COLORS["With_RAG"],
            ax=ax,
            inner="quartile",
            cut=0,
        )
    else:
        data = [df[df["model"] == m]["rubric_a_total"].dropna().tolist() for m in order]
        ax.boxplot(data, labels=order)
        ax.tick_params(axis="x", rotation=35)

    ax.set_xlabel("Model")
    ax.set_ylabel("Rubric A Total (/10)")
    ax.set_title("Rubric A Score Distribution by Model")
    ax.set_ylim(-0.5, 10.5)
    if HAS_SEABORN:
        ax.set_xticks(range(len(order)))
        ax.set_xticklabels(order, rotation=35, ha="right")
    fig.tight_layout()
    save_fig(fig, "rubric_a_distribution_by_model.png")


def fig_rubric_b_distribution(df: pd.DataFrame, order: list[str], compare_rag: bool) -> None:
    fig, ax = plt.subplots(figsize=(13, 6))

    if HAS_SEABORN and compare_rag:
        sns.violinplot(
            data=df,
            x="model",
            y="rubric_b_total",
            hue="rag_status",
            order=order,
            palette=RAG_COLORS,
            ax=ax,
            inner="quartile",
            cut=0,
        )
        handles, labels = ax.get_legend_handles_labels()
        ax.legend(handles, [RAG_LABELS.get(l, l) for l in labels])
    elif HAS_SEABORN:
        sns.violinplot(
            data=df,
            x="model",
            y="rubric_b_total",
            order=order,
            color=RAG_COLORS["With_RAG"],
            ax=ax,
            inner="quartile",
            cut=0,
        )
    else:
        data = [df[df["model"] == m]["rubric_b_total"].dropna().tolist() for m in order]
        ax.boxplot(data, labels=order)
        ax.tick_params(axis="x", rotation=35)

    ax.set_xlabel("Model")
    ax.set_ylabel("Rubric B Total (/6)")
    ax.set_title("Rubric B Score Distribution by Model")
    ax.set_ylim(-0.5, 6.5)
    if HAS_SEABORN:
        ax.set_xticks(range(len(order)))
        ax.set_xticklabels(order, rotation=35, ha="right")
    fig.tight_layout()
    save_fig(fig, "rubric_b_distribution_by_model.png")


def fig_heatmap(df: pd.DataFrame, order: list[str], value_col: str, filename: str, title: str, vmax: float) -> None:
    pivot = df.pivot_table(index="model", columns="task_id", values=value_col, aggfunc="mean")
    task_order = sorted(pivot.columns, key=lambda t: (int(t[1:]) if t[1:].isdigit() else 99, t))
    pivot = pivot.reindex(order).reindex(columns=task_order)

    fig, ax = plt.subplots(figsize=(14, 6))
    if HAS_SEABORN:
        sns.heatmap(
            pivot,
            ax=ax,
            cmap="RdYlGn",
            vmin=0,
            vmax=vmax,
            annot=True,
            fmt=".1f",
            linewidths=0.5,
            cbar_kws={"label": title},
        )
    else:
        im = ax.imshow(pivot.values, cmap="RdYlGn", aspect="auto", vmin=0, vmax=vmax)
        ax.set_xticks(range(len(task_order)))
        ax.set_xticklabels(task_order, rotation=45, ha="right")
        ax.set_yticks(range(len(order)))
        ax.set_yticklabels(order)
        fig.colorbar(im, ax=ax, label=title)

    ax.set_title(title)
    ax.set_xlabel("Task ID")
    ax.set_ylabel("Model")
    fig.tight_layout()
    save_fig(fig, filename)


def fig_tier_comparison(df: pd.DataFrame, order: list[str]) -> None:
    tier_agg = (
        df.groupby(["tier", "model"])["rubric_a_total"]
        .mean()
        .unstack()
        .reindex(TIER_ORDER)
        .reindex(columns=order)
    )

    fig, ax = plt.subplots(figsize=(12, 6))
    x = np.arange(len(TIER_ORDER))
    width = 0.15
    n_models = len(order)

    for i, model in enumerate(order):
        offset = (i - n_models / 2 + 0.5) * width
        vals = tier_agg[model].fillna(0).tolist()
        ax.bar(x + offset, vals, width, label=model, alpha=0.85)

    ax.set_xticks(x)
    ax.set_xticklabels([TIER_SHORT.get(t, t) for t in TIER_ORDER])
    ax.set_ylabel("Mean Rubric A (/10)")
    ax.set_title("Mean Rubric A by Task Tier and Model (RAG ON)")
    ax.set_ylim(0, 10.5)
    ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=8)
    ax.axhline(9, color="grey", linestyle="--", alpha=0.5)
    fig.tight_layout()
    save_fig(fig, "mean_rubric_a_by_tier.png")


def fig_sub_rubric_breakdown(df: pd.DataFrame, order: list[str]) -> None:
    a_cols = [c for c in SUB_RUBRIC_A if c in df.columns]
    b_cols = [c for c in SUB_RUBRIC_B if c in df.columns]

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    a_means = df.groupby("model")[a_cols].mean().reindex(order)
    b_means = df.groupby("model")[b_cols].mean().reindex(order)

    x_a = np.arange(len(a_cols))
    width = 0.15
    for i, model in enumerate(order):
        offset = (i - len(order) / 2 + 0.5) * width
        axes[0].bar(
            x_a + offset,
            a_means.loc[model].tolist(),
            width,
            label=model,
            alpha=0.85,
        )
    axes[0].set_xticks(x_a)
    axes[0].set_xticklabels([SUB_RUBRIC_A_LABELS[c] for c in a_cols], rotation=30, ha="right")
    axes[0].set_ylabel("Mean Score (/2)")
    axes[0].set_title("Rubric A Sub-criteria")
    axes[0].set_ylim(0, 2.2)

    x_b = np.arange(len(b_cols))
    for i, model in enumerate(order):
        offset = (i - len(order) / 2 + 0.5) * width
        axes[1].bar(
            x_b + offset,
            b_means.loc[model].tolist(),
            width,
            label=model,
            alpha=0.85,
        )
    axes[1].set_xticks(x_b)
    axes[1].set_xticklabels([SUB_RUBRIC_B_LABELS[c] for c in b_cols], rotation=30, ha="right")
    axes[1].set_ylabel("Mean Score (/2)")
    axes[1].set_title("Rubric B Sub-criteria")
    axes[1].set_ylim(0, 2.2)

    axes[1].legend(bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=8)
    fig.suptitle("Sub-rubric Breakdown by Model (RAG ON)", fontsize=12)
    fig.tight_layout()
    save_fig(fig, "sub_rubric_breakdown_by_model.png")


def fig_success_rate(summary: pd.DataFrame, order: list[str], compare_rag: bool) -> None:
    rag_vals, no_vals = _summary_values(
        summary,
        order,
        "success_rate_With_RAG",
        "success_rate_Without_RAG" if compare_rag else None,
    )
    rag_pct = [v * 100 for v in rag_vals]
    no_pct = [v * 100 for v in no_vals] if no_vals else None

    fig, ax = plt.subplots(figsize=(11, 5))
    grouped_bar(
        ax,
        order,
        rag_pct,
        no_pct,
        "Execution Success Rate (%)",
        "A1 Full Execution Success Rate by Model\n(A1 = 2: runs without manual correction)",
        ylim=(0, 110),
    )
    fig.tight_layout()
    save_fig(fig, "success_rate_by_model.png")


def fig_per_task_ranking(df: pd.DataFrame) -> None:
    tasks = sorted(df["task_id"].unique(), key=lambda t: (int(t[1:]) if t[1:].isdigit() else 99, t))
    n_tasks = len(tasks)

    ncols = 4
    nrows = int(np.ceil(n_tasks / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(4 * ncols, 3.5 * nrows), sharey=True)
    axes_flat = np.array(axes).flatten()

    for ax, task in zip(axes_flat, tasks):
        sub = df[df["task_id"] == task]
        grp = sub.groupby("model")["rubric_a_total"].mean().sort_values(ascending=True)
        ax.barh(grp.index, grp.values, color=RAG_COLORS["With_RAG"], alpha=0.8)
        task_name = sub["task_name"].iloc[0] if len(sub) else task
        short = task_name if len(task_name) <= 28 else task_name[:26] + "…"
        ax.set_title(f"{task}: {short}", fontsize=7)
        ax.set_xlim(0, 10.5)
        ax.set_xlabel("Rubric A")

    for ax in axes_flat[n_tasks:]:
        ax.set_visible(False)

    fig.suptitle("Model Ranking per Task (Rubric A, RAG ON)", fontsize=12)
    fig.tight_layout()
    save_fig(fig, "per_task_ranking.png")


def fig_rag_delta(summary: pd.DataFrame) -> None:
    col_a = "rubric_a_delta_RAG_minus_no_RAG"
    col_b = "rubric_b_delta_RAG_minus_no_RAG"
    if col_a not in summary.columns and col_b not in summary.columns:
        print("[INFO] RAG delta columns absent; skipping rag_delta figure.")
        return

    fig, axes = plt.subplots(1, 2, figsize=(13, 6))

    for ax, col, label, xlim in [
        (axes[0], col_a, "Δ Rubric A (With RAG − Without RAG)", (-5, 5)),
        (axes[1], col_b, "Δ Rubric B (With RAG − Without RAG)", (-3, 3)),
    ]:
        if col not in summary.columns:
            ax.set_visible(False)
            continue
        s = summary[["model", col]].dropna(subset=[col]).sort_values(col, ascending=True)
        colors = ["#4CAF50" if v >= 0 else "#E53935" for v in s[col]]
        ax.barh(s["model"], s[col], color=colors, alpha=0.85)
        ax.axvline(0, color="black", linewidth=1)
        ax.set_xlabel(label)
        ax.set_title(label)
        ax.set_xlim(*xlim)
        for i, v in enumerate(s[col]):
            ha = "left" if v >= 0 else "right"
            offset = 0.05 if v >= 0 else -0.05
            ax.text(v + offset, i, f"{v:+.2f}", va="center", ha=ha, fontsize=8)

    fig.suptitle("RAG Impact by Model", fontsize=12)
    fig.tight_layout()
    save_fig(fig, "rag_delta_by_model.png")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    args = parse_args(sys.argv[1:])
    print("\n" + "=" * 60)
    print("Urban EO/AI Evaluation Analysis")
    print("=" * 60)

    if args.input_xlsx:
        in_path = Path(args.input_xlsx)
        candidates = [
            in_path.resolve() if in_path.is_absolute() else None,
            (URBAN_DIR / in_path).resolve(),
            (URBAN_DIR / in_path.name).resolve(),
            (URBAN_DIR.parent.parent / in_path).resolve(),
        ]
        xlsx_path = next((p for p in candidates if p and p.exists()), None)
        if xlsx_path is None:
            sys.exit(f"ERROR: Input file does not exist: {args.input_xlsx}")
        print(f"[INFO] Using input file: {xlsx_path}")
    else:
        xlsx_path = find_excel_file(URBAN_DIR)

    df = load_experiment_runs(xlsx_path)
    validate_and_print(df)

    compare_rag = has_rag_comparison(df)

    summary = build_summary(df)
    build_tier_summary(df)
    build_task_summary(df)

    order = model_order(summary)

    print("\n[INFO] Generating figures...")
    fig_mean_rubric_a(summary, order, compare_rag)
    fig_mean_rubric_b(summary, order, compare_rag)
    fig_code_vs_interpretation(df)
    fig_rubric_a_distribution(df, order, compare_rag)
    fig_rubric_b_distribution(df, order, compare_rag)
    fig_heatmap(
        df,
        order,
        "rubric_a_total",
        "heatmap_rubric_a_by_task_and_model.png",
        "Rubric A Total (/10)",
        RUBRIC_A_MAX,
    )
    fig_heatmap(
        df,
        order,
        "rubric_b_total",
        "heatmap_rubric_b_by_task_and_model.png",
        "Rubric B Total (/6)",
        RUBRIC_B_MAX,
    )
    fig_tier_comparison(df, order)
    fig_sub_rubric_breakdown(df, order)
    fig_success_rate(summary, order, compare_rag)
    fig_per_task_ranking(df)
    if compare_rag:
        fig_rag_delta(summary)

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Input file      : {xlsx_path}")
    print(f"  Runs analysed   : {len(df)}")
    print(f"  Models          : {df['model'].nunique()}")
    for m in sorted(df["model"].unique()):
        print(f"                    • {m}")
    print(f"  Tasks           : {df['task_id'].nunique()}")
    print(f"  RAG comparison  : {'Yes' if compare_rag else 'No (ON only)'}")
    print(f"  Figures saved to: {RESULTS_DIR}")
    print(f"  Summary CSV     : {RESULTS_DIR / 'model_summary.csv'}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
