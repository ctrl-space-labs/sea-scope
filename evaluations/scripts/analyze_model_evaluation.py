"""
SeaScope LLM Remote-Sensing Evaluation Analysis
================================================
Reads the evaluation Excel workbook, validates data, produces a summary CSV,
and generates publication-ready figures for RAG vs Without-RAG comparison.

Usage:
    python evaluations/scripts/analyze_model_evaluation.py
"""

from __future__ import annotations

import argparse
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=FutureWarning)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
EVAL_DIR = REPO_ROOT / "evaluations"
RESULTS_DIR = REPO_ROOT / "evaluations" / "results"
SCRIPTS_DIR = REPO_ROOT / "evaluations" / "scripts"

RESULTS_DIR.mkdir(parents=True, exist_ok=True)
SCRIPTS_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Optional seaborn import
# ---------------------------------------------------------------------------

try:
    import seaborn as sns  # type: ignore

    HAS_SEABORN = True
    sns.set_theme(style="whitegrid", palette="muted")
except ImportError:
    HAS_SEABORN = False

import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

matplotlib.rcParams.update(
    {
        "figure.dpi": 150,
        "savefig.dpi": 150,
        "font.size": 10,
        "axes.titlesize": 12,
        "axes.labelsize": 11,
        "xtick.labelsize": 9,
        "ytick.labelsize": 9,
        "legend.fontsize": 9,
    }
)

# ---------------------------------------------------------------------------
# Colour palette (consistent across all figures)
# ---------------------------------------------------------------------------

RAG_COLORS = {"With_RAG": "#2196F3", "Without_RAG": "#FF7043"}
RAG_LABELS = {"With_RAG": "With RAG", "Without_RAG": "Without RAG"}

# ---------------------------------------------------------------------------
# Step 1 — Locate Excel file
# ---------------------------------------------------------------------------


def find_excel_file(directory: Path) -> Path:
    files = sorted(directory.glob("*.xlsx"))
    if not files:
        sys.exit(f"ERROR: No .xlsx file found in {directory}")
    chosen = files[0]
    if len(files) > 1:
        print(f"[INFO] Multiple .xlsx files found; using: {chosen.name}")
    else:
        print(f"[INFO] Using: {chosen.name}")
    return chosen


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="analyze_model_evaluation.py",
        description=(
            "Analyze LLM remote-sensing evaluations from an Excel workbook. "
            "Generates summary CSV + publication-ready figures comparing With_RAG vs Without_RAG."
        ),
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python evaluations/scripts/analyze_model_evaluation.py\n"
            "  python evaluations/scripts/analyze_model_evaluation.py --input evaluations/SeaScope\\ -\\ Model\\ Evaluation.xlsx\n"
        ),
    )
    parser.add_argument(
        "--input",
        "-i",
        dest="input_xlsx",
        type=str,
        default=None,
        help=(
            "Path to the input .xlsx workbook. If omitted, the script uses the first .xlsx "
            "file (sorted by filename) found in ./evaluations/."
        ),
    )
    return parser.parse_args(argv)


# ---------------------------------------------------------------------------
# Step 2 — Load and parse workbook
# ---------------------------------------------------------------------------

# Sheet-name → rag_status mapping (case-insensitive prefix match)
SHEET_RAG_MAP = {
    "rag": "With_RAG",
    "no rag": "Without_RAG",
    "without rag": "Without_RAG",
}


def sheet_to_rag_status(sheet_name: str) -> str:
    lower = sheet_name.strip().lower()
    for prefix, status in SHEET_RAG_MAP.items():
        if lower == prefix or lower.startswith(prefix):
            return status
    # Fallback: if "no" or "without" is anywhere in the name treat as Without_RAG
    if "no" in lower or "without" in lower:
        return "Without_RAG"
    return "With_RAG"


def load_workbook_data(xlsx_path: Path) -> pd.DataFrame:
    """Read all sheets, forward-fill case_study, tag rag_status."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        sys.exit(
            "ERROR: openpyxl is required. Install with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    frames: list[pd.DataFrame] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row is the header
        raw_header = rows[0]
        data_rows = rows[1:]

        # Build DataFrame; use range index for now
        df = pd.DataFrame(data_rows, columns=range(len(raw_header)))

        # Rename columns using normalised header names
        col_map = {}
        for i, h in enumerate(raw_header):
            if h is None:
                col_map[i] = f"_extra_{i}"
            else:
                normed = (
                    str(h)
                    .strip()
                    .lower()
                    .replace(" ", "_")
                    .replace("/", "_")
                    .replace("(", "")
                    .replace(")", "")
                )
                col_map[i] = normed
        df.rename(columns=col_map, inplace=True)

        # Drop entirely blank rows (all meaningful columns are None)
        meaningful = [c for c in df.columns if not c.startswith("_extra_")]
        df.dropna(subset=meaningful, how="all", inplace=True)

        # Forward-fill case_study (column 0 in original, likely 'case_study')
        case_col = [c for c in df.columns if "case" in c]
        if case_col:
            df[case_col[0]] = df[case_col[0]].ffill()

        # Drop trailing extra columns
        df = df[[c for c in df.columns if not c.startswith("_extra_")]]

        # Tag rag_status
        df["rag_status"] = sheet_to_rag_status(sheet_name)

        frames.append(df)
        print(f"  Sheet '{sheet_name}' → rag_status={sheet_to_rag_status(sheet_name)}, {len(df)} rows")

    wb.close()

    if not frames:
        sys.exit("ERROR: No data found in any sheet.")

    combined = pd.concat(frames, ignore_index=True)
    return combined


# ---------------------------------------------------------------------------
# Step 3 — Normalise column names and types
# ---------------------------------------------------------------------------

# Canonical column detection: keyword sets for each semantic role
COLUMN_ROLES = {
    "model": ["model"],
    "case_study": ["case_study", "case", "study", "scenario", "task"],
    "succeed_failed": ["succeed", "failed", "success", "pass"],
    "number_of_messages": ["message", "msg", "number_of_message"],
    "evaluation": ["evaluation", "eval", "score"],
    "notes": ["note"],
}


def detect_column(df: pd.DataFrame, role: str) -> str | None:
    """Return the first column name that matches any keyword for a given role."""
    keywords = COLUMN_ROLES.get(role, [])
    for col in df.columns:
        for kw in keywords:
            if kw in col:
                return col
    return None


def normalise_data(df: pd.DataFrame) -> pd.DataFrame:
    role_to_col: dict[str, str | None] = {
        role: detect_column(df, role) for role in COLUMN_ROLES
    }

    print("\n[INFO] Detected columns:")
    for role, col in role_to_col.items():
        print(f"  {role:25s} → {col}")

    for required in ["model", "evaluation", "number_of_messages"]:
        if role_to_col[required] is None:
            sys.exit(f"ERROR: Cannot detect required column for role '{required}'")

    # Rename to canonical names
    rename = {}
    for role, col in role_to_col.items():
        if col and col != role:
            rename[col] = role
    df = df.rename(columns=rename)

    # Cast numeric columns; '×' strings become NaN
    df["evaluation"] = pd.to_numeric(df["evaluation"], errors="coerce")
    df["number_of_messages"] = pd.to_numeric(df["number_of_messages"], errors="coerce")

    # Derive boolean success from succeed_failed symbol
    if "succeed_failed" in df.columns:
        df["success"] = df["succeed_failed"].astype(str).str.strip() == "✓"
    else:
        # Fallback: infer from evaluation score
        threshold = 5
        print(
            f"[WARN] No succeed/failed column found. "
            f"Inferring success as evaluation >= {threshold}."
        )
        df["success"] = df["evaluation"] >= threshold

    # Ensure rag_status is present
    if "rag_status" not in df.columns:
        sys.exit("ERROR: rag_status column missing after loading.")

    # Drop rows where both model and evaluation are missing
    df = df.dropna(subset=["model", "evaluation"], how="all")

    # Strip whitespace from model names
    df["model"] = df["model"].astype(str).str.strip()

    return df


# ---------------------------------------------------------------------------
# Step 4 — Validation diagnostics
# ---------------------------------------------------------------------------


def validate_and_print(df: pd.DataFrame) -> None:
    print("\n" + "=" * 60)
    print("DATA VALIDATION")
    print("=" * 60)

    print(f"\nColumns and dtypes:\n{df.dtypes}\n")

    print("Row counts by model and rag_status:")
    counts = df.groupby(["model", "rag_status"]).size().unstack(fill_value=0)
    print(counts.to_string())

    print("\n\nRaw evaluation values per model and rag_status:")
    for (model, rag), grp in df.groupby(["model", "rag_status"]):
        vals = grp["evaluation"].dropna().tolist()
        s = grp["evaluation"].dropna().sum()
        n = grp["evaluation"].dropna().count()
        mean = grp["evaluation"].mean()
        print(
            f"  {model:30s} | {rag:12s} | n={n:2d} | sum={s:5.1f} | "
            f"mean={mean:5.2f} | values={vals}"
        )

    print("\nIdentical-mean check:")
    summary = (
        df.groupby(["model", "rag_status"])["evaluation"]
        .mean()
        .unstack()
    )
    for model in summary.index:
        rag_mean = summary.loc[model].get("With_RAG", None)
        no_rag_mean = summary.loc[model].get("Without_RAG", None)
        if rag_mean is not None and no_rag_mean is not None:
            if abs(rag_mean - no_rag_mean) < 1e-9:
                print(
                    f"  [!] {model}: identical means ({rag_mean:.4f}) — "
                    f"verifying raw values..."
                )
                for rag in ["With_RAG", "Without_RAG"]:
                    grp = df[(df["model"] == model) & (df["rag_status"] == rag)]
                    vals = grp["evaluation"].dropna().tolist()
                    print(f"       {rag}: n={len(vals)} sum={sum(vals):.1f} values={vals}")

    print("=" * 60)


# ---------------------------------------------------------------------------
# Step 5 — Build model_summary.csv
# ---------------------------------------------------------------------------


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    grp = df.groupby(["model", "rag_status"])

    agg = grp.agg(
        mean_eval=("evaluation", "mean"),
        median_eval=("evaluation", "median"),
        std_eval=("evaluation", "std"),
        mean_msg=("number_of_messages", "mean"),
        success_rate=("success", "mean"),
        n=("evaluation", "count"),
    ).reset_index()

    # Pivot: one row per model, columns split by rag_status
    def pivot_metric(metric: str) -> pd.DataFrame:
        return agg.pivot(index="model", columns="rag_status", values=metric)

    metrics = ["mean_eval", "median_eval", "std_eval", "mean_msg", "success_rate"]
    pivoted = [pivot_metric(m) for m in metrics]

    summary = pd.concat(pivoted, axis=1)
    # Flatten multi-level column names
    summary.columns = [f"{m}_{rag}" for m in metrics for rag in pivoted[0].columns]
    summary = summary.reset_index()

    # Compute deltas (With_RAG − Without_RAG)
    for metric in ["mean_eval", "mean_msg"]:
        col_rag = f"{metric}_With_RAG"
        col_no = f"{metric}_Without_RAG"
        delta_col = f"{metric.replace('mean_', '')}_delta_RAG_minus_no_RAG"
        if col_rag in summary.columns and col_no in summary.columns:
            summary[delta_col] = summary[col_rag] - summary[col_no]

    csv_path = RESULTS_DIR / "model_summary.csv"
    summary.to_csv(csv_path, index=False, float_format="%.4f")
    print(f"\n[OK] model_summary.csv → {csv_path}")
    return summary


# ---------------------------------------------------------------------------
# Shared helpers for figures
# ---------------------------------------------------------------------------


def save_fig(fig: plt.Figure, name: str) -> None:
    path = RESULTS_DIR / name
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    print(f"[OK] {name} → {path}")


def model_order(summary: pd.DataFrame) -> list[str]:
    """Sort models by descending mean eval With_RAG (or Without_RAG fallback)."""
    col = "mean_eval_With_RAG" if "mean_eval_With_RAG" in summary.columns else summary.columns[1]
    return summary.sort_values(col, ascending=False)["model"].tolist()


def grouped_bar(
    ax: plt.Axes,
    models: list[str],
    rag_values: list[float],
    no_rag_values: list[float],
    ylabel: str,
    title: str,
    ylim: tuple | None = None,
) -> None:
    x = np.arange(len(models))
    width = 0.35
    ax.bar(x - width / 2, rag_values, width, label="With RAG", color=RAG_COLORS["With_RAG"], alpha=0.85)
    ax.bar(x + width / 2, no_rag_values, width, label="Without RAG", color=RAG_COLORS["Without_RAG"], alpha=0.85)
    ax.set_xticks(x)
    ax.set_xticklabels(models, rotation=40, ha="right")
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.legend()
    if ylim:
        ax.set_ylim(*ylim)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=False, nbins=6))


# ---------------------------------------------------------------------------
# Figure A — Overall mean evaluation by model
# ---------------------------------------------------------------------------


def fig_A_overall_mean_eval(summary: pd.DataFrame, order: list[str]) -> None:
    s = summary.set_index("model").reindex(order)
    fig, ax = plt.subplots(figsize=(12, 5))
    grouped_bar(
        ax,
        order,
        s.get("mean_eval_With_RAG", pd.Series()).fillna(0).tolist(),
        s.get("mean_eval_Without_RAG", pd.Series()).fillna(0).tolist(),
        "Mean Evaluation Score (/10)",
        "Overall Mean Evaluation Score by Model",
        ylim=(0, 10.5),
    )
    ax.axhline(5, color="grey", linestyle="--", linewidth=0.8, label="Score = 5")
    ax.legend()
    fig.tight_layout()
    save_fig(fig, "overall_mean_evaluation_by_model.png")


# ---------------------------------------------------------------------------
# Figure B — RAG evaluation delta
# ---------------------------------------------------------------------------


def fig_B_rag_delta(summary: pd.DataFrame, order: list[str]) -> None:
    """Horizontal bar chart sorted by delta (largest at top), coloured by model-size tier."""
    from matplotlib.patches import Patch

    col = "eval_delta_RAG_minus_no_RAG"
    s = summary.set_index("model")
    if col not in s.columns:
        print("[WARN] Delta column missing; skipping figure B.")
        return

    # Sort: largest delta at top (ascending=True → top of barh is last item)
    s_sorted = s[[col]].dropna().sort_values(col, ascending=True)
    models_sorted = s_sorted.index.tolist()
    deltas_sorted = s_sorted[col].tolist()

    bar_colors = [TIER_COLORS.get(TIER_MAP.get(m, "Compact"), "#888888") for m in models_sorted]

    fig, ax = plt.subplots(figsize=(11, 8))

    # Background sign zones use green/red only for RAG effect direction (not tier colour).
    x_min = min(deltas_sorted) - 0.4
    x_max = max(deltas_sorted) + 0.4
    ax.axvspan(0, x_max, color="#E8F5E9", alpha=0.35, zorder=0)
    ax.axvspan(x_min, 0, color="#FFEBEE", alpha=0.35, zorder=0)

    ax.barh(range(len(models_sorted)), deltas_sorted, color=bar_colors, alpha=0.92, zorder=2, edgecolor="white", linewidth=0.6)
    ax.axvline(0, color="black", linewidth=1.4, zorder=3)

    # Value labels on bars
    for i, v in enumerate(deltas_sorted):
        ha = "left" if v >= 0 else "right"
        offset = 0.05 if v >= 0 else -0.05
        ax.text(v + offset, i, f"{v:+.2f}", va="center", ha=ha, fontsize=8.5, fontweight="bold")

    ax.set_yticks(range(len(models_sorted)))
    ax.set_yticklabels(models_sorted, fontsize=9)
    ax.set_xlabel("Δ Evaluation Score (With RAG − Without RAG)", fontsize=10)
    ax.set_title(
        "RAG Impact by Model — Who Benefits, Who Struggles?\n"
        "(sorted by impact; bar colour = model size tier)",
        fontsize=11,
        pad=12,
    )

    ax.text(x_max * 0.98, len(models_sorted) - 0.4, "RAG helps", ha="right", va="top",
            fontsize=9, color="#2E7D32", style="italic")
    ax.text(x_min * 0.98, -0.35, "RAG hurts", ha="left", va="bottom",
            fontsize=9, color="#C62828", style="italic")

    legend_els = [
        Patch(facecolor=TIER_COLORS[t], label=f"{t} models")
        for t in ["Frontier", "Mid-size", "Compact"]
    ]
    ax.legend(handles=legend_els, loc="lower right", fontsize=9, title="Model size tier")

    ax.grid(axis="x", linestyle="--", alpha=0.4, zorder=1)
    fig.tight_layout()
    save_fig(fig, "rag_evaluation_delta_by_model.png")


# ---------------------------------------------------------------------------
# Figure C — Mean messages by model
# ---------------------------------------------------------------------------


def fig_C_mean_messages(summary: pd.DataFrame, order: list[str]) -> None:
    s = summary.set_index("model").reindex(order)
    fig, ax = plt.subplots(figsize=(12, 5))
    grouped_bar(
        ax,
        order,
        s.get("mean_msg_With_RAG", pd.Series()).fillna(0).tolist(),
        s.get("mean_msg_Without_RAG", pd.Series()).fillna(0).tolist(),
        "Mean Number of Messages",
        "Mean Number of Messages by Model",
    )
    # Red cut-off line at 10 messages (observed convergence / failure threshold in the data)
    ax.axhline(10, color="#E53935", linestyle="--", linewidth=1.8, zorder=3)
    ax.text(
        len(order) - 0.5, 10.25,
        "10-message cut-off",
        ha="right", va="bottom",
        color="#E53935", fontsize=8.5, fontweight="bold",
    )
    fig.tight_layout()
    save_fig(fig, "mean_messages_by_model.png")


# ---------------------------------------------------------------------------
# Figure D — Success rate by model
# ---------------------------------------------------------------------------


def fig_D_success_rate(summary: pd.DataFrame, order: list[str]) -> None:
    s = summary.set_index("model").reindex(order)
    fig, ax = plt.subplots(figsize=(12, 5))
    grouped_bar(
        ax,
        order,
        (s.get("success_rate_With_RAG", pd.Series()).fillna(0) * 100).tolist(),
        (s.get("success_rate_Without_RAG", pd.Series()).fillna(0) * 100).tolist(),
        "Success Rate (%)",
        "Success Rate by Model (Succeeded / Total)",
        ylim=(0, 110),
    )
    fig.tight_layout()
    save_fig(fig, "success_rate_by_model.png")


# ---------------------------------------------------------------------------
# Figure E — Efficiency vs quality scatter
# ---------------------------------------------------------------------------


def _shorten_model_name(name: str) -> str:
    """
    Human-readable short label that keeps enough words to disambiguate.
    E.g. "Claude Opus 4.6" → "Claude Opus 4.6" (kept in full when ≤ 3 words)
         "Gemini 3.1 Flash-light" → "Gemini Flash-light"
         "OpenAI GPT-5-mini"      → "GPT-5-mini"
    """
    parts = name.split()
    if len(parts) <= 3:
        return name          # already short enough
    # Drop purely numeric version tokens in the middle to shorten
    filtered = [p for p in parts if not p.replace(".", "").isdigit()]
    return " ".join(filtered) if filtered else name


def _repel_labels(points: list[tuple[float, float]],
                  offsets: list[tuple[float, float]],
                  x_range: float,
                  y_range: float,
                  iterations: int = 40) -> list[tuple[float, float]]:
    """
    Minimal force-directed repulsion for annotation offsets (in data units).
    Keeps labels from sitting exactly on top of each other.
    """
    import copy
    off = [list(o) for o in offsets]
    min_dx = x_range * 0.04
    min_dy = y_range * 0.06

    for _ in range(iterations):
        for i in range(len(off)):
            lx = points[i][0] + off[i][0]
            ly = points[i][1] + off[i][1]
            for j in range(len(off)):
                if i == j:
                    continue
                lx2 = points[j][0] + off[j][0]
                ly2 = points[j][1] + off[j][1]
                ddx = lx - lx2
                ddy = ly - ly2
                if abs(ddx) < min_dx and abs(ddy) < min_dy:
                    push_x = min_dx - abs(ddx)
                    push_y = min_dy - abs(ddy)
                    sign_x = 1 if ddx >= 0 else -1
                    sign_y = 1 if ddy >= 0 else -1
                    off[i][0] += sign_x * push_x * 0.5
                    off[i][1] += sign_y * push_y * 0.5
    return [tuple(o) for o in off]


def fig_E_efficiency_vs_quality(df: pd.DataFrame) -> None:
    try:
        from adjustText import adjust_text  # type: ignore
        HAS_ADJUSTTEXT = True
    except ImportError:
        HAS_ADJUSTTEXT = False

    grp = (
        df.groupby(["model", "rag_status"])
        .agg(mean_eval=("evaluation", "mean"), mean_msg=("number_of_messages", "mean"))
        .reset_index()
    )

    fig, ax = plt.subplots(figsize=(14, 9))

    # Circle = With RAG, triangle-up = Without RAG
    MARKERS = {"With_RAG": "o", "Without_RAG": "^"}

    texts = []
    for rag in ["With_RAG", "Without_RAG"]:
        sub = grp[grp["rag_status"] == rag]
        ax.scatter(
            sub["mean_msg"],
            sub["mean_eval"],
            label=RAG_LABELS[rag],
            color=RAG_COLORS[rag],
            marker=MARKERS[rag],
            s=110,
            zorder=3,
            alpha=0.90,
            edgecolors="white",
            linewidths=0.8,
        )
        rag_tag = "RAG" if rag == "With_RAG" else "No RAG"
        for _, row in sub.iterrows():
            short = _shorten_model_name(row["model"])
            t = ax.text(
                row["mean_msg"],
                row["mean_eval"],
                f"{short} ({rag_tag})",
                fontsize=7.5,
                color="dimgray",
                bbox=dict(boxstyle="round,pad=0.2", facecolor="white",
                          edgecolor="lightgray", alpha=0.85, linewidth=0.5),
            )
            texts.append(t)

    if HAS_ADJUSTTEXT:
        adjust_text(
            texts,
            ax=ax,
            arrowprops=dict(arrowstyle="-", color="lightgray", lw=0.7),
            expand=(1.4, 1.6),
            force_text=(0.6, 0.8),
            force_points=(0.3, 0.4),
            min_arrow_len=4,
        )
    else:
        # Fallback: nudge texts away from their anchor points
        for t in texts:
            t.set_position((t.get_position()[0] + 0.1, t.get_position()[1] + 0.15))

    ax.set_xlabel("Mean Number of Messages (lower = more efficient)", fontsize=10)
    ax.set_ylabel("Mean Evaluation Score (/10)", fontsize=10)
    ax.set_title("Efficiency vs Quality: Mean Messages vs Mean Evaluation Score", fontsize=11)

    # Legend: RAG colour + marker shape
    from matplotlib.lines import Line2D
    legend_els = [
        Line2D([0], [0], marker="o", color="w", markerfacecolor=RAG_COLORS["With_RAG"],
               markersize=9, label="With RAG  (●)"),
        Line2D([0], [0], marker="^", color="w", markerfacecolor=RAG_COLORS["Without_RAG"],
               markersize=9, label="Without RAG  (▲)"),
    ]
    ax.legend(handles=legend_els, fontsize=9)
    ax.set_ylim(-0.5, 11.5)
    ax.grid(linestyle="--", alpha=0.35)
    fig.tight_layout()
    save_fig(fig, "efficiency_vs_quality.png")


# ---------------------------------------------------------------------------
# Figure F — Evaluation distribution (boxplot/violin)
# ---------------------------------------------------------------------------


def fig_F_eval_distribution(df: pd.DataFrame, order: list[str]) -> None:
    fig, ax = plt.subplots(figsize=(14, 6))

    if HAS_SEABORN:
        sns.violinplot(
            data=df,
            x="model",
            y="evaluation",
            hue="rag_status",
            order=order,
            palette=RAG_COLORS,
            ax=ax,
            split=False,
            inner="quartile",
            cut=0,
        )
    else:
        # Manual grouped boxplot via matplotlib
        positions_rag, positions_no = [], []
        data_rag, data_no = [], []
        x_ticks, x_labels = [], []

        for i, model in enumerate(order):
            base = i * 3
            grp_rag = df[(df["model"] == model) & (df["rag_status"] == "With_RAG")]["evaluation"].dropna()
            grp_no = df[(df["model"] == model) & (df["rag_status"] == "Without_RAG")]["evaluation"].dropna()
            positions_rag.append(base)
            positions_no.append(base + 1)
            data_rag.append(grp_rag.tolist())
            data_no.append(grp_no.tolist())
            x_ticks.append(base + 0.5)
            x_labels.append(model)

        bp1 = ax.boxplot(data_rag, positions=positions_rag, widths=0.7, patch_artist=True,
                         boxprops=dict(facecolor=RAG_COLORS["With_RAG"], alpha=0.7),
                         medianprops=dict(color="white", linewidth=2))
        bp2 = ax.boxplot(data_no, positions=positions_no, widths=0.7, patch_artist=True,
                         boxprops=dict(facecolor=RAG_COLORS["Without_RAG"], alpha=0.7),
                         medianprops=dict(color="white", linewidth=2))
        ax.set_xticks(x_ticks)
        ax.set_xticklabels(x_labels, rotation=40, ha="right")

        from matplotlib.patches import Patch
        ax.legend(handles=[
            Patch(facecolor=RAG_COLORS["With_RAG"], label="With RAG"),
            Patch(facecolor=RAG_COLORS["Without_RAG"], label="Without RAG"),
        ])

    ax.set_xlabel("Model")
    ax.set_ylabel("Evaluation Score (/10)")
    ax.set_title("Evaluation Score Distribution by Model and RAG Status")
    ax.set_ylim(-0.5, 11)

    if HAS_SEABORN:
        # Fix x-tick labels after seaborn (must set_ticks before set_xticklabels)
        ax.set_xticks(range(len(order)))
        ax.set_xticklabels(order, rotation=40, ha="right")
        handles, labels = ax.get_legend_handles_labels()
        ax.legend(handles, [RAG_LABELS.get(l, l) for l in labels])

    fig.tight_layout()
    save_fig(fig, "evaluation_distribution_by_model.png")


# ---------------------------------------------------------------------------
# Figure G — RAG delta heatmap by case study and model
# ---------------------------------------------------------------------------


def fig_G_rag_delta_heatmap(df: pd.DataFrame, order: list[str]) -> None:
    if "case_study" not in df.columns:
        print("[INFO] No case_study column; skipping figure G.")
        return

    pivot_rag = df[df["rag_status"] == "With_RAG"].pivot_table(
        index="model", columns="case_study", values="evaluation", aggfunc="mean"
    )
    pivot_no = df[df["rag_status"] == "Without_RAG"].pivot_table(
        index="model", columns="case_study", values="evaluation", aggfunc="mean"
    )
    delta = (pivot_rag - pivot_no).reindex(order)

    if delta.empty:
        print("[WARN] Delta heatmap is empty; skipping figure G.")
        return

    fig, ax = plt.subplots(figsize=(14, 7))
    if HAS_SEABORN:
        sns.heatmap(
            delta,
            ax=ax,
            cmap="RdBu",
            center=0,
            annot=True,
            fmt=".1f",
            linewidths=0.5,
            cbar_kws={"label": "Δ Score (With RAG − Without RAG)"},
        )
    else:
        im = ax.imshow(delta.values, cmap="RdBu", aspect="auto", vmin=-5, vmax=5)
        ax.set_xticks(range(len(delta.columns)))
        ax.set_xticklabels(delta.columns, rotation=45, ha="right")
        ax.set_yticks(range(len(delta.index)))
        ax.set_yticklabels(delta.index)
        for i in range(len(delta.index)):
            for j in range(len(delta.columns)):
                val = delta.values[i, j]
                if not np.isnan(val):
                    ax.text(j, i, f"{val:.1f}", ha="center", va="center", fontsize=8)
        fig.colorbar(im, ax=ax, label="Δ Score (With RAG − Without RAG)")

    ax.set_title("RAG Evaluation Delta by Case Study and Model")
    ax.set_xlabel("Case Study")
    ax.set_ylabel("Model")
    fig.tight_layout()
    save_fig(fig, "rag_delta_heatmap_by_case_and_model.png")


# ---------------------------------------------------------------------------
# Figure H — Raw eval heatmap (RAG / No RAG side by side)
# ---------------------------------------------------------------------------


def fig_H_eval_heatmap(df: pd.DataFrame, order: list[str]) -> None:
    """Side-by-side heatmaps (stacked vertically) — green=best, red=worst."""
    if "case_study" not in df.columns:
        print("[INFO] No case_study column; skipping figure H.")
        return

    cases = sorted(df["case_study"].dropna().unique())
    # Stack vertically (2 rows, 1 col) so the x-axis labels have plenty of horizontal room
    fig, axes = plt.subplots(2, 1, figsize=(14, 16), sharex=True)

    for ax, rag in zip(axes, ["With_RAG", "Without_RAG"]):
        pivot = (
            df[df["rag_status"] == rag]
            .pivot_table(index="model", columns="case_study", values="evaluation", aggfunc="mean")
            .reindex(order)
            .reindex(columns=cases)
        )

        if HAS_SEABORN:
            sns.heatmap(
                pivot,
                ax=ax,
                cmap="RdYlGn",   # red=0 (worst) → yellow → green=10 (best)
                vmin=0,
                vmax=10,
                annot=True,
                fmt=".1f",
                linewidths=0.5,
                annot_kws={"size": 9},
                cbar_kws={"label": "Mean Eval Score (/10)", "shrink": 0.6},
            )
        else:
            im = ax.imshow(pivot.values, cmap="RdYlGn", aspect="auto", vmin=0, vmax=10)
            ax.set_xticks(range(len(cases)))
            ax.set_xticklabels(cases, rotation=45, ha="right", fontsize=9)
            ax.set_yticks(range(len(order)))
            ax.set_yticklabels(order, fontsize=9)
            for i in range(len(order)):
                for j in range(len(cases)):
                    val = pivot.values[i, j]
                    if not np.isnan(val):
                        ax.text(j, i, f"{val:.1f}", ha="center", va="center", fontsize=8)
            fig.colorbar(im, ax=ax, label="Mean Eval Score (/10)", shrink=0.6)

        ax.set_title(f"Evaluation Scores — {RAG_LABELS[rag]}", fontsize=11, pad=8)
        ax.set_ylabel("Model", fontsize=10)
        if ax is axes[-1]:
            ax.set_xlabel("Case Study", fontsize=10)
            ax.tick_params(axis="x", rotation=45, labelsize=9)
        ax.tick_params(axis="y", labelsize=9)

    fig.suptitle("Evaluation Score Heatmap per Model × Case Study\n(green = best, red = worst)", fontsize=13)
    fig.tight_layout(rect=[0, 0, 1, 0.97])
    save_fig(fig, "evaluation_heatmap_by_case_and_model.png")


# ---------------------------------------------------------------------------
# Figure I — Model family comparison
# ---------------------------------------------------------------------------

FAMILY_MAP = {
    "Gemini 3.1 PRO": "Gemini",
    "Gemini 3 Flash": "Gemini",
    "Gemini 3.1 Flash-light": "Gemini",
    "OpenAI GPT-5.1": "OpenAI",
    "OpenAI GPT-5-mini": "OpenAI",
    "OpenAI GPT-5-nano": "OpenAI",
    "Claude Opus 4.6": "Claude",
    "Claude Sonnet 4.6": "Claude",
    "Claude Haiku 4.5": "Claude",
    "GPT OSS 20B": "OSS",
    "GPT OSS 120B": "OSS",
    "Qwen3 32B": "OSS",
    "Llama 4 Scout": "OSS",
}

# Model size tier (editorial grouping for storytelling). Tier colours are distinct
# from the green/red “RAG helps / hurts” background zones.
TIER_MAP = {
    "Gemini 3.1 PRO": "Frontier",
    "Claude Opus 4.6": "Frontier",
    "Gemini 3 Flash": "Frontier",
    "OpenAI GPT-5.1": "Frontier",
    "Gemini 3.1 Flash-light": "Mid-size",
    "Claude Sonnet 4.6": "Mid-size",
    "OpenAI GPT-5-mini": "Mid-size",
    "Claude Haiku 4.5": "Mid-size",
    "GPT OSS 120B": "Mid-size",
    "OpenAI GPT-5-nano": "Compact",
    "GPT OSS 20B": "Compact",
    "Qwen3 32B": "Compact",
    "Llama 4 Scout": "Compact",
}

TIER_COLORS = {
    "Frontier": "#42A5F4",
    "Mid-size": "#FF855F",
    "Compact":  "#8271e3",
}


def fig_I_model_family(df: pd.DataFrame) -> None:
    df2 = df.copy()
    df2["family"] = df2["model"].map(FAMILY_MAP).fillna("Other")

    family_agg = (
        df2.groupby(["family", "rag_status"])["evaluation"]
        .mean()
        .unstack()
        .reindex(columns=["With_RAG", "Without_RAG"])
    )
    families = family_agg.index.tolist()

    fig, ax = plt.subplots(figsize=(9, 5))
    grouped_bar(
        ax,
        families,
        family_agg.get("With_RAG", pd.Series()).fillna(0).tolist(),
        family_agg.get("Without_RAG", pd.Series()).fillna(0).tolist(),
        "Mean Evaluation Score (/10)",
        "Mean Evaluation Score by Model Family",
        ylim=(0, 10.5),
    )
    ax.axhline(5, color="grey", linestyle="--", linewidth=0.8)
    fig.tight_layout()
    save_fig(fig, "model_family_comparison.png")


# ---------------------------------------------------------------------------
# Figure J — Per-case-study model ranking
# ---------------------------------------------------------------------------


def fig_J_per_case_ranking(df: pd.DataFrame) -> None:
    if "case_study" not in df.columns:
        print("[INFO] No case_study column; skipping figure J.")
        return

    cases = sorted(df["case_study"].dropna().unique())
    n_cases = len(cases)

    fig, axes = plt.subplots(1, n_cases, figsize=(4 * n_cases, 7), sharey=True)
    if n_cases == 1:
        axes = [axes]

    for ax, case in zip(axes, cases):
        sub = df[df["case_study"] == case]
        for rag, color in RAG_COLORS.items():
            grp = (
                sub[sub["rag_status"] == rag]
                .groupby("model")["evaluation"]
                .mean()
                .sort_values(ascending=True)
            )
            ax.barh(
                grp.index,
                grp.values,
                color=color,
                alpha=0.7,
                label=RAG_LABELS[rag],
            )
        short = case if len(case) <= 20 else case[:18] + "…"
        ax.set_title(short, fontsize=8)
        ax.set_xlim(0, 10.5)
        if ax == axes[0]:
            ax.set_ylabel("Model")
        ax.set_xlabel("Score")
        ax.tick_params(axis="y", labelsize=7)

    handles, labels = axes[0].get_legend_handles_labels()
    # Deduplicate legend labels
    seen = {}
    for h, l in zip(handles, labels):
        seen.setdefault(l, h)
    fig.legend(seen.values(), seen.keys(), loc="upper right", fontsize=8)
    fig.suptitle("Model Ranking per Case Study", fontsize=12)
    fig.tight_layout()
    save_fig(fig, "per_case_study_ranking.png")


# ---------------------------------------------------------------------------
# Figure K — Success vs eval score scatter (individual rows)
# ---------------------------------------------------------------------------


def fig_K_success_vs_eval(df: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(9, 5))
    rng = np.random.default_rng(42)

    for rag, color in RAG_COLORS.items():
        sub = df[df["rag_status"] == rag].dropna(subset=["evaluation", "success"])
        jitter_y = rng.uniform(-0.06, 0.06, size=len(sub))
        ax.scatter(
            sub["evaluation"],
            sub["success"].astype(int) + jitter_y,
            color=color,
            alpha=0.55,
            s=35,
            label=RAG_LABELS[rag],
        )

    ax.set_xlabel("Evaluation Score (/10)")
    ax.set_ylabel("Outcome")
    ax.set_yticks([0, 1])
    ax.set_yticklabels(["Failed", "Succeeded"])
    ax.set_title("Individual Evaluation Score vs Task Outcome (with jitter)")
    ax.legend()
    ax.set_xlim(-0.5, 10.5)
    fig.tight_layout()
    save_fig(fig, "success_vs_eval_score_scatter.png")


# ---------------------------------------------------------------------------
# Figure L — RAG Sweet Spot: baseline capability vs RAG impact
# ---------------------------------------------------------------------------


def fig_L_rag_sweet_spot(df: pd.DataFrame) -> None:
    """
    Scatter plot: x = model's WITHOUT-RAG mean score (baseline capability),
    y = RAG delta (With_RAG − Without_RAG).

    Point colour = model size tier (Frontier / Mid-size / Compact).
    """
    grp = (
        df.groupby(["model", "rag_status"])["evaluation"]
        .mean()
        .unstack()
        .rename(columns={"With_RAG": "rag", "Without_RAG": "no_rag"})
    )
    grp["delta"] = grp["rag"] - grp["no_rag"]
    grp = grp.dropna(subset=["no_rag", "delta"]).reset_index()
    grp["tier"] = grp["model"].map(TIER_MAP).fillna("Compact")

    fig, ax = plt.subplots(figsize=(11, 7))

    for tier, color in TIER_COLORS.items():
        sub = grp[grp["tier"] == tier]
        ax.scatter(sub["no_rag"], sub["delta"], color=color, s=120, zorder=3,
                   label=f"{tier} models", alpha=0.92, edgecolors="white", linewidths=0.9)
        for _, row in sub.iterrows():
            # Keep Claude Opus/Sonnet/Haiku distinguishable
            short = _shorten_model_name(row["model"])
            ax.annotate(
                short,
                (row["no_rag"], row["delta"]),
                textcoords="offset points",
                xytext=(7, 3),
                fontsize=8,
                color=color,
            )

    # Horizontal zero line
    ax.axhline(0, color="black", linewidth=1.1, linestyle="--", alpha=0.6)

    # Soft background zones to guide the reader's eye
    ylim_lo, ylim_hi = ax.get_ylim()
    ax.axhspan(0, ylim_hi + 1, color="#E8F5E9", alpha=0.35, zorder=0)   # positive zone
    ax.axhspan(ylim_lo - 1, 0, color="#FFEBEE", alpha=0.35, zorder=0)   # negative zone

    ax.text(0.5, 0.97, "RAG helps", transform=ax.transAxes,
            ha="center", va="top", fontsize=9, color="#2E7D32", style="italic")
    ax.text(0.5, 0.03, "RAG hurts", transform=ax.transAxes,
            ha="center", va="bottom", fontsize=9, color="#C62828", style="italic")

    ax.set_xlabel("Baseline Score Without RAG (higher = stronger model)", fontsize=10)
    ax.set_ylabel("RAG Impact: Δ Score (With RAG − Without RAG)", fontsize=10)
    ax.set_title(
        "The RAG Sweet Spot: Which Models Benefit Most from Retrieval-Augmented Context?\n"
        "Point colour = model size tier (teal / coral / purple)",
        fontsize=10.5, pad=12,
    )
    ax.legend(fontsize=9, title="Model size tier", loc="upper left")
    ax.grid(linestyle="--", alpha=0.35, zorder=1)

    fig.tight_layout()
    save_fig(fig, "rag_sweet_spot_baseline_vs_delta.png")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    args = parse_args(sys.argv[1:])
    print("\n" + "=" * 60)
    print("SeaScope LLM Evaluation Analysis")
    print("=" * 60)

    # 1 — Find file
    if args.input_xlsx:
        in_path = Path(args.input_xlsx)
        xlsx_path = (REPO_ROOT / in_path).resolve() if not in_path.is_absolute() else in_path
        if not xlsx_path.exists():
            sys.exit(f"ERROR: Input file does not exist: {xlsx_path}")
        if xlsx_path.suffix.lower() != ".xlsx":
            sys.exit(f"ERROR: Input file must be a .xlsx workbook: {xlsx_path}")
        print(f"[INFO] Using input file from --input: {xlsx_path}")
    else:
        xlsx_path = find_excel_file(EVAL_DIR)

    # 2 — Load
    print("\n[INFO] Loading sheets...")
    df_raw = load_workbook_data(xlsx_path)

    # 3 — Normalise
    df = normalise_data(df_raw)

    # 4 — Validate
    validate_and_print(df)

    # 5 — Summary CSV
    summary = build_summary(df)

    # Model order for consistent x-axis
    order = model_order(summary)

    # 6 — Figures
    print("\n[INFO] Generating figures...")
    fig_A_overall_mean_eval(summary, order)
    fig_B_rag_delta(summary, order)
    fig_C_mean_messages(summary, order)
    fig_D_success_rate(summary, order)
    fig_E_efficiency_vs_quality(df)
    fig_F_eval_distribution(df, order)
    fig_G_rag_delta_heatmap(df, order)
    fig_H_eval_heatmap(df, order)
    fig_I_model_family(df)
    fig_J_per_case_ranking(df)
    fig_K_success_vs_eval(df)
    fig_L_rag_sweet_spot(df)

    # 7 — Terminal summary
    models_found = sorted(df["model"].unique())
    rag_statuses = df["rag_status"].unique().tolist()

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Input file      : {xlsx_path}")
    print(f"  Rows analysed   : {len(df)}")
    print(f"  Models found    : {len(models_found)}")
    for m in models_found:
        print(f"                    • {m}")
    print(f"  RAG statuses    : {', '.join(rag_statuses)}")
    print(f"  With_RAG        : {'Yes' if 'With_RAG' in rag_statuses else 'No'}")
    print(f"  Without_RAG     : {'Yes' if 'Without_RAG' in rag_statuses else 'No'}")
    print(f"  Figures saved to: {RESULTS_DIR}")
    print(f"  Summary CSV     : {RESULTS_DIR / 'model_summary.csv'}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
