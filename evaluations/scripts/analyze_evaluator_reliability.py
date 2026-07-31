"""
SeaScope Evaluator Reliability & Model Consistency Analysis
============================================================
Reads the v2 evaluation workbook (which adds a 2nd evaluator's 5 repeated
runs for a single model, on top of the original 1st-evaluator single-run
data), and produces:

  1. Internal consistency ("test-retest") statistics: how much a model's
     score/behaviour varies across 5 independent repeated runs of the same
     case study, scored by Evaluator 2.
  2. Inter-evaluator overlap statistics: how well Evaluator 1's single score
     agrees with Evaluator 2's mean-of-5 score, per case study & RAG
     condition (ICC, correlation, bias, Bland-Altman).

Data shape assumed (see evaluations/REPORT.md for full context):
  - Sheets "RAG" / "NO RAG", each with a two-row header.
  - Columns A-F ("base" block): Evaluator 1's original single run
    (Case study, Model, Succeed/Failed, Number of messages, Evaluation, Notes).
  - Columns grouped under "1st eveluation" .. "5th eveluation": Evaluator 2's
    5 independent repeated runs (Succeed/Failed, Number of messages,
    Evaluation), populated only for the target model (default: Gemini 3 Flash).
  - A trailing "Average of the 5 evaluations" block, which we use only as a
    sanity cross-check against our own recomputed averages.

Usage:
    python evaluations/scripts/analyze_evaluator_reliability.py
    python evaluations/scripts/analyze_evaluator_reliability.py --model "Gemini 3 Flash"
"""

from __future__ import annotations

import argparse
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from scipy import stats as sstats

warnings.filterwarnings("ignore", category=FutureWarning)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
EVAL_DIR = REPO_ROOT / "evaluations"
RESULTS_DIR = REPO_ROOT / "evaluations" / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_MODEL = "Gemini 3 Flash"

# ---------------------------------------------------------------------------
# Plot styling (kept consistent with analyze_model_evaluation.py)
# ---------------------------------------------------------------------------

try:
    import seaborn as sns  # type: ignore

    HAS_SEABORN = True
    sns.set_theme(style="whitegrid", palette="muted")
except ImportError:
    HAS_SEABORN = False

import matplotlib
import matplotlib.pyplot as plt

matplotlib.rcParams.update(
    {
        "figure.dpi": 150,
        "savefig.dpi": 150,
        "font.size": 10,
        "axes.titlesize": 12,
        "axes.labelsize": 11,
        "xtick.labelsize": 9,
        "ytick.labelsize": 9,
        "legend.fontsize": 9,
    }
)

EXPORT_DPI = 300

RAG_COLORS = {"With_RAG": "#2196F3", "Without_RAG": "#FF7043"}
RAG_LABELS = {"With_RAG": "With RAG", "Without_RAG": "Without RAG"}


def save_fig(fig: plt.Figure, name: str) -> None:
    path = RESULTS_DIR / name
    fig.savefig(path, dpi=EXPORT_DPI, bbox_inches="tight", pad_inches=0.05,
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"[OK] {name} → {path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="analyze_evaluator_reliability.py",
        description=(
            "Analyze inter-evaluator overlap and within-model (test-retest) "
            "consistency from the v2 evaluation workbook."
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--input", "-i", dest="input_xlsx", type=str, default=None,
        help=(
            "Path to the v2 .xlsx workbook. If omitted, the script looks for "
            "the first .xlsx file in ./evaluations/ whose name contains 'v2'."
        ),
    )
    parser.add_argument(
        "--model", "-m", dest="model", type=str, default=DEFAULT_MODEL,
        help=f"Model name to analyze (default: '{DEFAULT_MODEL}').",
    )
    return parser.parse_args(argv)


def find_v2_excel_file(directory: Path) -> Path:
    candidates = sorted(p for p in directory.glob("*.xlsx") if "v2" in p.stem.lower())
    if not candidates:
        sys.exit(
            f"ERROR: No 'v2' .xlsx file found in {directory}. "
            "Pass one explicitly with --input."
        )
    chosen = candidates[0]
    if len(candidates) > 1:
        print(f"[INFO] Multiple 'v2' .xlsx files found; using: {chosen.name}")
    else:
        print(f"[INFO] Using: {chosen.name}")
    return chosen


# ---------------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------------

SHEET_RAG_MAP = {"rag": "With_RAG", "no rag": "Without_RAG", "without rag": "Without_RAG"}


def sheet_to_rag_status(sheet_name: str) -> str:
    lower = sheet_name.strip().lower()
    for prefix, status in SHEET_RAG_MAP.items():
        if lower == prefix or lower.startswith(prefix):
            return status
    if "no" in lower or "without" in lower:
        return "Without_RAG"
    return "With_RAG"


FIELD_KEYWORDS = {
    "case_study": ["case"],
    "model": ["model"],
    "succeed_failed": ["succeed", "failed"],
    "number_of_messages": ["message"],
    "evaluation": ["evaluation"],
    "notes": ["note"],
}

EVAL_ROUND_LABELS = [
    "1st eveluation", "2nd eveluation", "3rd eveluation",
    "4th eveluation", "5th eveluation",
]
AVERAGE_LABEL = "average of the 5 evaluations"


def _normalize_field(header_value) -> str | None:
    if header_value is None:
        return None
    low = str(header_value).strip().lower()
    for field, keywords in FIELD_KEYWORDS.items():
        if any(kw in low for kw in keywords):
            return field
    return None


def parse_header_blocks(row0: tuple, row1: tuple) -> dict[str, dict[str, int]]:
    """Map each 'group' (base / Nth evaluation / average) to {field: col_idx}."""
    blocks: dict[str, dict[str, int]] = {}
    current_group = "base"
    for i, (group_label, field_label) in enumerate(zip(row0, row1)):
        if group_label is not None:
            current_group = str(group_label).strip().lower()
        field = _normalize_field(field_label)
        if field is None:
            continue
        blocks.setdefault(current_group, {})[field] = i
    return blocks


def load_v2_workbook(xlsx_path: Path, target_model: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      long_df: one row per (case_study, rag_status, source, ...) where
               source in {"evaluator_1", "evaluator_2_run_1".."evaluator_2_run_5"}
               — restricted to `target_model` only.
      sheet_avg_df: the workbook's own precomputed "Average of the 5
               evaluations" values, kept separately for a sanity cross-check.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        sys.exit("ERROR: openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    records: list[dict] = []
    avg_records: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        row0, row1, data_rows = rows[0], rows[1], rows[2:]
        blocks = parse_header_blocks(row0, row1)
        rag_status = sheet_to_rag_status(sheet_name)

        if "base" not in blocks or "case_study" not in blocks["base"] or "model" not in blocks["base"]:
            sys.exit(f"ERROR: Could not detect base 'case_study'/'model' columns in sheet '{sheet_name}'.")

        base = blocks["base"]
        case_col, model_col = base["case_study"], base["model"]

        current_case = None
        n_model_rows = 0
        for r in data_rows:
            if all(v is None for v in r):
                continue
            if r[case_col] is not None:
                current_case = r[case_col]
            model_val = r[model_col]
            if model_val is None or str(model_val).strip() != target_model:
                continue

            n_model_rows += 1
            case_study = current_case

            # Evaluator 1 — base block
            records.append({
                "case_study": case_study,
                "rag_status": rag_status,
                "source": "evaluator_1",
                "succeed_failed": r[base["succeed_failed"]] if "succeed_failed" in base else None,
                "number_of_messages": r[base["number_of_messages"]] if "number_of_messages" in base else None,
                "evaluation": r[base["evaluation"]] if "evaluation" in base else None,
            })

            # Evaluator 2 — 5 repeated runs
            for run_idx, label in enumerate(EVAL_ROUND_LABELS, start=1):
                if label not in blocks:
                    continue
                b = blocks[label]
                records.append({
                    "case_study": case_study,
                    "rag_status": rag_status,
                    "source": f"evaluator_2_run_{run_idx}",
                    "succeed_failed": r[b["succeed_failed"]] if "succeed_failed" in b else None,
                    "number_of_messages": r[b["number_of_messages"]] if "number_of_messages" in b else None,
                    "evaluation": r[b["evaluation"]] if "evaluation" in b else None,
                })

            # Workbook's own precomputed average (cross-check only)
            if AVERAGE_LABEL in blocks:
                ab = blocks[AVERAGE_LABEL]
                avg_records.append({
                    "case_study": case_study,
                    "rag_status": rag_status,
                    "sheet_avg_messages": r[ab["number_of_messages"]] if "number_of_messages" in ab else None,
                    "sheet_avg_evaluation": r[ab["evaluation"]] if "evaluation" in ab else None,
                })

        print(f"  Sheet '{sheet_name}' → rag_status={rag_status}, {n_model_rows} row(s) matched model='{target_model}'")

    wb.close()

    if not records:
        sys.exit(f"ERROR: No rows found for model='{target_model}'. Check --model spelling.")

    long_df = pd.DataFrame(records)
    long_df["evaluation"] = pd.to_numeric(long_df["evaluation"], errors="coerce")
    long_df["number_of_messages"] = pd.to_numeric(long_df["number_of_messages"], errors="coerce")
    # Keep missing outcomes missing. Treating a blank cell as False would turn
    # incomplete workbook data into an apparent model failure.
    long_df["success"] = (
        long_df["succeed_failed"]
        .astype("string")
        .str.strip()
        .map({"✓": True, "×": False})
        .astype("boolean")
    )

    sheet_avg_df = pd.DataFrame(avg_records)
    if not sheet_avg_df.empty:
        sheet_avg_df["sheet_avg_messages"] = pd.to_numeric(sheet_avg_df["sheet_avg_messages"], errors="coerce")
        sheet_avg_df["sheet_avg_evaluation"] = pd.to_numeric(sheet_avg_df["sheet_avg_evaluation"], errors="coerce")

    return long_df, sheet_avg_df


# ---------------------------------------------------------------------------
# Statistics helpers
# ---------------------------------------------------------------------------


def safe_cv(std: float, mean: float) -> float:
    return std / mean if mean not in (0, None) and not np.isnan(mean) and mean != 0 else np.nan


def compute_icc(ratings: np.ndarray) -> dict[str, float]:
    """
    Shrout & Fleiss (1979) / McGraw & Wong (1996) two-way ANOVA ICC formulas.
    ratings: shape (n_targets, k_raters).
    Returns ICC(2,1) [absolute agreement] and ICC(3,1) [consistency].
    """
    ratings = np.asarray(ratings, dtype=float)
    n, k = ratings.shape
    grand_mean = ratings.mean()
    row_means = ratings.mean(axis=1)
    col_means = ratings.mean(axis=0)

    ss_total = np.sum((ratings - grand_mean) ** 2)
    ss_rows = k * np.sum((row_means - grand_mean) ** 2)
    ss_cols = n * np.sum((col_means - grand_mean) ** 2)
    ss_error = ss_total - ss_rows - ss_cols

    ms_rows = ss_rows / (n - 1)
    ms_cols = ss_cols / (k - 1)
    ms_error = ss_error / ((n - 1) * (k - 1))

    icc2_1 = (ms_rows - ms_error) / (ms_rows + (k - 1) * ms_error + k * (ms_cols - ms_error) / n)
    icc3_1 = (ms_rows - ms_error) / (ms_rows + (k - 1) * ms_error)

    return {"icc2_1": float(icc2_1), "icc3_1": float(icc3_1)}


# ---------------------------------------------------------------------------
# Validation / cross-check
# ---------------------------------------------------------------------------


def validate_and_print(long_df: pd.DataFrame, sheet_avg_df: pd.DataFrame, target_model: str) -> None:
    print("\n" + "=" * 70)
    print(f"DATA VALIDATION — model='{target_model}'")
    print("=" * 70)

    e2 = long_df[long_df["source"].str.startswith("evaluator_2")]
    complete = e2[
        e2["evaluation"].notna()
        & e2["number_of_messages"].notna()
        & e2["success"].notna()
    ]
    counts = complete.groupby(["case_study", "rag_status"]).size()
    n_groups = counts.shape[0]
    bad = counts[counts != 5]
    print(f"\nGroups with Evaluator-2 data: {n_groups} (expected 14 = 7 case studies x 2 conditions)")
    expected_groups = 14
    if len(bad) or n_groups != expected_groups:
        print("[ERROR] Incomplete Evaluator-2 data; every group must contain 5 scores, message counts, and outcomes:")
        print(bad.to_string())
        sys.exit(1)
    else:
        print("[OK] Every group has exactly 5 complete Evaluator-2 runs.")

    if sheet_avg_df.empty:
        print("[INFO] No 'Average of the 5 evaluations' block found; skipping cross-check.")
    else:
        recomputed = (
            e2.groupby(["case_study", "rag_status"])
            .agg(recomputed_avg_eval=("evaluation", "mean"), recomputed_avg_msg=("number_of_messages", "mean"))
            .reset_index()
        )
        merged = recomputed.merge(sheet_avg_df, on=["case_study", "rag_status"], how="left")
        merged["eval_diff"] = (merged["recomputed_avg_eval"] - merged["sheet_avg_evaluation"]).abs()
        merged["msg_diff"] = (merged["recomputed_avg_msg"] - merged["sheet_avg_messages"]).abs()
        tolerance = 0.05
        mismatches = merged[
            merged["sheet_avg_evaluation"].isna()
            | merged["sheet_avg_messages"].isna()
            | (merged["eval_diff"] > tolerance)
            | (merged["msg_diff"] > tolerance)
        ]
        print(f"\nCross-check vs. workbook's own 'Average of the 5 evaluations' (tolerance={tolerance}):")
        if mismatches.empty:
            print("[OK] Recomputed 5-run averages match the workbook's precomputed averages.")
        else:
            print("[WARN] Mismatches found:")
            print(mismatches.to_string(index=False))

    print("=" * 70)


# ---------------------------------------------------------------------------
# Section A — Internal consistency (within Evaluator 2's 5 runs)
# ---------------------------------------------------------------------------


def build_group_table(long_df: pd.DataFrame) -> pd.DataFrame:
    e1 = long_df[long_df["source"] == "evaluator_1"]
    e2 = long_df[long_df["source"].str.startswith("evaluator_2")]

    case_order = long_df.drop_duplicates("case_study")["case_study"].tolist()

    rows = []
    for (case_study, rag_status), sub in e2.groupby(["case_study", "rag_status"], sort=False):
        e1_row = e1[(e1["case_study"] == case_study) & (e1["rag_status"] == rag_status)]
        if e1_row.empty:
            continue
        e1_eval = float(e1_row["evaluation"].iloc[0])
        e1_msg = float(e1_row["number_of_messages"].iloc[0])
        e1_success = bool(e1_row["success"].iloc[0])

        evals = sub["evaluation"].dropna().to_numpy()
        msgs = sub["number_of_messages"].dropna().to_numpy()
        succs = sub["success"].dropna().astype(bool).to_numpy()

        if len(evals) != 5 or len(msgs) != 5 or len(succs) != 5:
            sys.exit(
                f"ERROR: Expected 5 complete repeated runs for "
                f"case_study='{case_study}', rag_status='{rag_status}'."
            )

        mean2, median2, std2 = evals.mean(), float(np.median(evals)), evals.std(ddof=1)
        # Round the 5-run mean to the nearest whole score for inter-evaluator
        # comparison (matches the integer 0–10 scale used by Evaluator 1).
        mean2_rounded = float(np.rint(mean2))
        min2, max2 = evals.min(), evals.max()
        iqr2 = float(np.percentile(evals, 75) - np.percentile(evals, 25))
        msg_mean2, msg_std2 = msgs.mean(), msgs.std(ddof=1)
        succ_rate2 = float(succs.mean())

        bias = e1_eval - mean2_rounded
        within_range = bool(min2 <= e1_eval <= max2)
        within_1sd = bool(abs(e1_eval - mean2) <= std2) if std2 > 0 else bool(e1_eval == mean2)

        rows.append({
            "case_study": case_study,
            "rag_status": rag_status,
            "evaluator1_evaluation": e1_eval,
            "evaluator1_messages": e1_msg,
            "evaluator1_success": e1_success,
            "evaluator2_n": len(evals),
            "evaluator2_mean_eval": mean2,
            "evaluator2_mean_eval_rounded": mean2_rounded,
            "evaluator2_median_eval": median2,
            "evaluator2_std_eval": std2,
            "evaluator2_cv_eval": safe_cv(std2, mean2),
            "evaluator2_min_eval": float(min2),
            "evaluator2_max_eval": float(max2),
            "evaluator2_range_eval": float(max2 - min2),
            "evaluator2_iqr_eval": iqr2,
            "evaluator2_mean_msg": msg_mean2,
            "evaluator2_std_msg": msg_std2,
            "evaluator2_cv_msg": safe_cv(msg_std2, msg_mean2),
            "evaluator2_success_rate": succ_rate2,
            "evaluator2_majority_success": succ_rate2 > 0.5,
            "bias_eval1_minus_eval2mean": bias,
            "abs_bias": abs(bias),
            "within_eval2_range": within_range,
            "within_1sd": within_1sd,
        })

    group_df = pd.DataFrame(rows)
    group_df["case_study"] = pd.Categorical(group_df["case_study"], categories=case_order, ordered=True)
    group_df = group_df.sort_values(["case_study", "rag_status"]).reset_index(drop=True)
    return group_df


# ---------------------------------------------------------------------------
# Section B — Inter-evaluator overlap (Evaluator 1 vs Evaluator 2 mean-of-5)
# ---------------------------------------------------------------------------


def build_overall_table(group_df: pd.DataFrame) -> pd.DataFrame:
    metrics: list[tuple[str, object, str]] = []

    def add(name, value, note=""):
        metrics.append((name, value, note))

    n_groups = len(group_df)
    add("n_groups", n_groups, "case_study x rag_status pairs")
    add("n_with_rag", int((group_df["rag_status"] == "With_RAG").sum()))
    add("n_without_rag", int((group_df["rag_status"] == "Without_RAG").sum()))

    # --- Internal consistency (pooled) ---
    for label, sub in [("overall", group_df),
                        ("with_rag", group_df[group_df["rag_status"] == "With_RAG"]),
                        ("without_rag", group_df[group_df["rag_status"] == "Without_RAG"])]:
        add(f"mean_within_group_sd_eval_{label}", sub["evaluator2_std_eval"].mean(),
            "avg of the 5-run std within each group — headline test-retest SD")
        add(f"mean_within_group_cv_eval_{label}", sub["evaluator2_cv_eval"].mean())
        add(f"mean_within_group_sd_msg_{label}", sub["evaluator2_std_msg"].mean())
        add(f"mean_within_group_range_eval_{label}", sub["evaluator2_range_eval"].mean())

    # Paired RAG-vs-no-RAG comparison across the same 7 case studies.
    paired = group_df.pivot(
        index="case_study",
        columns="rag_status",
        values=["evaluator2_mean_eval", "evaluator2_std_eval", "evaluator2_cv_eval"],
    )
    for metric, short_name in [
        ("evaluator2_mean_eval", "mean_eval"),
        ("evaluator2_std_eval", "within_group_sd_eval"),
        ("evaluator2_cv_eval", "within_group_cv_eval"),
    ]:
        with_rag = paired[(metric, "With_RAG")].to_numpy(dtype=float)
        without_rag = paired[(metric, "Without_RAG")].to_numpy(dtype=float)
        delta = with_rag - without_rag
        t_result = sstats.ttest_rel(with_rag, without_rag)
        w_result = sstats.wilcoxon(with_rag, without_rag, zero_method="pratt")
        add(f"{short_name}_with_rag", float(with_rag.mean()))
        add(f"{short_name}_without_rag", float(without_rag.mean()))
        add(f"{short_name}_delta_rag_minus_no_rag", float(delta.mean()))
        add(f"{short_name}_paired_ttest_p", float(t_result.pvalue))
        add(f"{short_name}_wilcoxon_p", float(w_result.pvalue))

    # --- Inter-evaluator overlap: Evaluator 1 vs rounded Evaluator 2 mean-of-5 ---
    e1 = group_df["evaluator1_evaluation"].to_numpy(dtype=float)
    e2m = group_df["evaluator2_mean_eval_rounded"].to_numpy(dtype=float)

    icc = compute_icc(np.column_stack([e1, e2m]))
    add("icc_2_1", icc["icc2_1"], "two-way random, absolute agreement — headline inter-rater reliability (E2 mean rounded)")
    add("icc_3_1", icc["icc3_1"], "two-way mixed, consistency-only (E2 mean rounded)")

    pearson_r, pearson_p = sstats.pearsonr(e1, e2m)
    add("pearson_r", pearson_r)
    add("pearson_p", pearson_p)

    spearman_rho, spearman_p = sstats.spearmanr(e1, e2m)
    add("spearman_rho", spearman_rho)
    add("spearman_p", spearman_p)

    diffs = e1 - e2m
    add("mean_bias_eval1_minus_eval2mean", float(diffs.mean()), "Evaluator1 − rounded Evaluator2 mean")
    add("std_bias", float(diffs.std(ddof=1)))
    add("bland_altman_upper_loa", float(diffs.mean() + 1.96 * diffs.std(ddof=1)))
    add("bland_altman_lower_loa", float(diffs.mean() - 1.96 * diffs.std(ddof=1)))

    try:
        t_stat, t_p = sstats.ttest_rel(e1, e2m)
        add("paired_ttest_stat", float(t_stat))
        add("paired_ttest_p", float(t_p))
    except Exception as exc:  # pragma: no cover
        add("paired_ttest_stat", np.nan, str(exc))
        add("paired_ttest_p", np.nan)

    try:
        w_stat, w_p = sstats.wilcoxon(e1, e2m, zero_method="pratt")
        add("wilcoxon_stat", float(w_stat))
        add("wilcoxon_p", float(w_p))
    except Exception as exc:  # pragma: no cover
        add("wilcoxon_stat", np.nan, str(exc))
        add("wilcoxon_p", np.nan)

    add("pct_within_eval2_range", float(group_df["within_eval2_range"].mean()))
    add("pct_within_1sd", float(group_df["within_1sd"].mean()))

    return pd.DataFrame(metrics, columns=["metric", "value", "note"])


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------


def fig_repeated_runs_by_case_study(long_df: pd.DataFrame, group_df: pd.DataFrame) -> None:
    e2 = long_df[long_df["source"].str.startswith("evaluator_2")]
    case_order = group_df["case_study"].cat.categories.tolist()

    fig, ax = plt.subplots(figsize=(12, 6))
    rng = np.random.default_rng(42)

    positions_rag, positions_no, x_ticks, x_labels = [], [], [], []
    data_rag, data_no = [], []

    for i, case in enumerate(case_order):
        base = i * 3
        positions_rag.append(base)
        positions_no.append(base + 1)
        x_ticks.append(base + 0.5)
        x_labels.append(case)
        for rag, data_store in [("With_RAG", data_rag), ("Without_RAG", data_no)]:
            vals = e2[(e2["case_study"] == case) & (e2["rag_status"] == rag)]["evaluation"].dropna().to_numpy()
            data_store.append(vals)

    ax.boxplot(
        data_rag,
        positions=positions_rag,
        widths=0.7,
        patch_artist=True,
        showfliers=False,
        boxprops=dict(facecolor=RAG_COLORS["With_RAG"], alpha=0.35),
        medianprops=dict(color=RAG_COLORS["With_RAG"], linewidth=2),
        zorder=2,
    )
    ax.boxplot(
        data_no,
        positions=positions_no,
        widths=0.7,
        patch_artist=True,
        showfliers=False,
        boxprops=dict(facecolor=RAG_COLORS["Without_RAG"], alpha=0.35),
        medianprops=dict(color=RAG_COLORS["Without_RAG"], linewidth=2),
        zorder=2,
    )

    for i, case in enumerate(case_order):
        for rag, pos in [("With_RAG", positions_rag[i]), ("Without_RAG", positions_no[i])]:
            vals = e2[(e2["case_study"] == case) & (e2["rag_status"] == rag)]["evaluation"].dropna().to_numpy()
            jitter = rng.uniform(-0.18, 0.18, size=len(vals))
            ax.scatter(np.full(len(vals), pos) + jitter, vals, color=RAG_COLORS[rag],
                       alpha=0.75, s=28, zorder=3, edgecolors="white", linewidths=0.4)

    from matplotlib.lines import Line2D
    legend_els = [
        Line2D([0], [0], marker="o", color="w", markerfacecolor=RAG_COLORS["With_RAG"], markersize=8, label="With RAG"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor=RAG_COLORS["Without_RAG"], markersize=8, label="Without RAG"),
    ]
    ax.legend(handles=legend_els, loc="lower left", fontsize=8.5)

    ax.set_xticks(x_ticks)
    ax.set_xticklabels(x_labels, rotation=30, ha="right")
    ax.set_ylabel("Evaluation Score (/10)")
    ax.set_ylim(-0.5, 10.8)
    ax.set_title("Gemini 3 Flash — Five-Run Consistency by Case Study")
    fig.tight_layout()
    save_fig(fig, "evaluator_repeated_runs_by_case_study.png")


def fig_rag_vs_norag_summary(group_df: pd.DataFrame) -> None:
    rag_summary = (
        group_df.groupby("rag_status", observed=True)
        .agg(
            mean_score=("evaluator2_mean_eval", "mean"),
            mean_case_sd=("evaluator2_std_eval", "mean"),
        )
        .reindex(["With_RAG", "Without_RAG"])
    )
    summary_colors = [RAG_COLORS["With_RAG"], RAG_COLORS["Without_RAG"]]
    summary_labels = ["With RAG", "Without RAG"]

    fig, axes = plt.subplots(1, 2, figsize=(9, 4.5))
    for ax, column, title, ylabel, ylim in [
        (axes[0], "mean_score", "Mean Evaluation Score", "Score (/10)", (0, 10)),
        (axes[1], "mean_case_sd", "Mean Within-Case SD", "SD (lower = steadier)", (0, 2)),
    ]:
        values = rag_summary[column].to_numpy()
        bars = ax.bar(summary_labels, values, color=summary_colors, alpha=0.85, width=0.55)
        ax.bar_label(bars, fmt="%.2f", padding=3, fontsize=10)
        ax.set_title(title, fontsize=11)
        ax.set_ylabel(ylabel, fontsize=10)
        ax.set_ylim(*ylim)
        ax.grid(axis="y", linestyle="--", alpha=0.3)

    fig.suptitle("Gemini 3 Flash — Five-Run Summary: RAG vs Without RAG", fontsize=12, y=1.02)
    fig.tight_layout()
    save_fig(fig, "evaluator_rag_vs_norag_summary.png")


def fig_bland_altman(group_df: pd.DataFrame, overall_df: pd.DataFrame) -> None:
    e1 = group_df["evaluator1_evaluation"].to_numpy(dtype=float)
    e2m = group_df["evaluator2_mean_eval_rounded"].to_numpy(dtype=float)
    means = (e1 + e2m) / 2
    diffs = e1 - e2m

    mean_bias = float(overall_df.set_index("metric").loc["mean_bias_eval1_minus_eval2mean", "value"])
    upper_loa = float(overall_df.set_index("metric").loc["bland_altman_upper_loa", "value"])
    lower_loa = float(overall_df.set_index("metric").loc["bland_altman_lower_loa", "value"])

    fig, ax = plt.subplots(figsize=(9, 6))
    rng = np.random.default_rng(7)
    for rag in ["With_RAG", "Without_RAG"]:
        mask = (group_df["rag_status"] == rag).to_numpy()
        n = int(mask.sum())
        jitter_x = rng.uniform(-0.08, 0.08, size=n)
        jitter_y = rng.uniform(-0.08, 0.08, size=n)
        ax.scatter(
            means[mask] + jitter_x,
            diffs[mask] + jitter_y,
            color=RAG_COLORS[rag],
            label=f"{RAG_LABELS[rag]} (n={n})",
            s=90,
            alpha=0.85,
            edgecolors="white",
            linewidths=0.6,
            zorder=3,
        )

    # Bland–Altman 95% limits of agreement: mean_diff ± 1.96 * SD(diff)
    ax.axhline(mean_bias, color="black", linewidth=1.4,
               label=f"Mean difference (bias): {mean_bias:+.2f}")
    ax.axhline(upper_loa, color="grey", linestyle="--", linewidth=1.2,
               label=f"Upper 95% LoA: {upper_loa:+.2f}")
    ax.axhline(lower_loa, color="grey", linestyle="--", linewidth=1.2,
               label=f"Lower 95% LoA: {lower_loa:+.2f}")
    ax.axhline(0, color="#B0BEC5", linewidth=1.0, linestyle=":")

    ax.set_xlabel("Mean of Evaluator 1 and Evaluator 2 scores")
    ax.set_ylabel("Difference (Evaluator 1 − Evaluator 2)")
    ax.set_title("Bland–Altman Agreement Plot — Evaluator 1 vs Evaluator 2 (Gemini 3 Flash)")
    ax.legend(fontsize=8.5, loc="best", title="95% LoA = bias ± 1.96·SD(diff)")
    ax.grid(linestyle="--", alpha=0.35)
    fig.tight_layout()
    save_fig(fig, "evaluator_bland_altman.png")


def fig_agreement_scatter(group_df: pd.DataFrame, overall_df: pd.DataFrame) -> None:
    ov = overall_df.set_index("metric")["value"]
    icc2_1 = float(ov.loc["icc_2_1"])
    pearson_r = float(ov.loc["pearson_r"])
    spearman_rho = float(ov.loc["spearman_rho"])

    fig, ax = plt.subplots(figsize=(8, 8))
    # Rounded integer scores often collide on the same (x, y); small jitter
    # keeps all 14 points visible without changing the underlying values.
    rng = np.random.default_rng(42)
    for rag in ["With_RAG", "Without_RAG"]:
        sub = group_df[group_df["rag_status"] == rag]
        x = sub["evaluator2_mean_eval_rounded"].to_numpy(dtype=float)
        y = sub["evaluator1_evaluation"].to_numpy(dtype=float)
        jitter_x = rng.uniform(-0.12, 0.12, size=len(x))
        jitter_y = rng.uniform(-0.12, 0.12, size=len(y))
        ax.scatter(
            x + jitter_x,
            y + jitter_y,
            color=RAG_COLORS[rag],
            label=f"{RAG_LABELS[rag]} (n={len(sub)})",
            s=100,
            alpha=0.85,
            edgecolors="white",
            linewidths=0.7,
            zorder=3,
        )

    lims = [-0.5, 10.5]
    ax.plot(lims, lims, color="grey", linestyle="--", linewidth=1.2, zorder=1, label="y = x (perfect agreement)")
    ax.set_xlim(lims)
    ax.set_ylim(lims)

    ax.text(
        0.03, 0.97,
        f"ICC(2,1) = {icc2_1:.2f}\nPearson r = {pearson_r:.2f}\nSpearman ρ = {spearman_rho:.2f}",
        transform=ax.transAxes, ha="left", va="top", fontsize=10,
        bbox=dict(boxstyle="round,pad=0.4", facecolor="white", edgecolor="lightgray", alpha=0.9),
    )

    ax.set_xlabel("Evaluator 2")
    ax.set_ylabel("Evaluator 1")
    ax.set_title("Inter-Evaluator Agreement — Gemini 3 Flash")
    ax.legend(fontsize=9, loc="lower right")
    ax.grid(linestyle="--", alpha=0.35)
    fig.tight_layout()
    save_fig(fig, "evaluator_agreement_scatter.png")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    args = parse_args(sys.argv[1:])
    print("\n" + "=" * 70)
    print("SeaScope Evaluator Reliability & Model Consistency Analysis")
    print("=" * 70)

    if args.input_xlsx:
        in_path = Path(args.input_xlsx)
        xlsx_path = (REPO_ROOT / in_path).resolve() if not in_path.is_absolute() else in_path
        if not xlsx_path.exists():
            sys.exit(f"ERROR: Input file does not exist: {xlsx_path}")
    else:
        xlsx_path = find_v2_excel_file(EVAL_DIR)

    print(f"\n[INFO] Loading sheets for model='{args.model}'...")
    long_df, sheet_avg_df = load_v2_workbook(xlsx_path, args.model)

    validate_and_print(long_df, sheet_avg_df, args.model)

    group_df = build_group_table(long_df)
    group_csv = RESULTS_DIR / "evaluator_reliability_by_group.csv"
    group_df.to_csv(group_csv, index=False, float_format="%.4f")
    print(f"\n[OK] evaluator_reliability_by_group.csv → {group_csv}")

    overall_df = build_overall_table(group_df)
    overall_csv = RESULTS_DIR / "evaluator_reliability_overall.csv"
    overall_df.to_csv(overall_csv, index=False)
    print(f"[OK] evaluator_reliability_overall.csv → {overall_csv}")

    print("\n[INFO] Generating figures...")
    fig_repeated_runs_by_case_study(long_df, group_df)
    fig_rag_vs_norag_summary(group_df)
    fig_bland_altman(group_df, overall_df)
    fig_agreement_scatter(group_df, overall_df)

    print("\n" + "=" * 70)
    print("HEADLINE NUMBERS")
    print("=" * 70)
    ov = overall_df.set_index("metric")["value"]
    print(f"  Groups analyzed              : {int(ov['n_groups'])} (7 case studies x 2 RAG conditions)")
    print(f"  Mean within-group SD (eval)  : {ov['mean_within_group_sd_eval_overall']:.3f}  "
          f"[With RAG: {ov['mean_within_group_sd_eval_with_rag']:.3f} | Without RAG: {ov['mean_within_group_sd_eval_without_rag']:.3f}]")
    print(f"  Mean within-group CV (eval)  : {ov['mean_within_group_cv_eval_overall']:.3f}")
    print(f"  Mean score, RAG vs no RAG    : {ov['mean_eval_with_rag']:.3f} vs {ov['mean_eval_without_rag']:.3f}  "
          f"(delta={ov['mean_eval_delta_rag_minus_no_rag']:+.3f}, Wilcoxon p={ov['mean_eval_wilcoxon_p']:.3f})")
    print(f"  Mean SD, RAG vs no RAG       : {ov['within_group_sd_eval_with_rag']:.3f} vs "
          f"{ov['within_group_sd_eval_without_rag']:.3f}  "
          f"(delta={ov['within_group_sd_eval_delta_rag_minus_no_rag']:+.3f}, "
          f"Wilcoxon p={ov['within_group_sd_eval_wilcoxon_p']:.3f})")
    print(f"  ICC(2,1) absolute agreement  : {ov['icc_2_1']:.3f}")
    print(f"  ICC(3,1) consistency         : {ov['icc_3_1']:.3f}")
    print(f"  Pearson r / Spearman rho     : {ov['pearson_r']:.3f} / {ov['spearman_rho']:.3f}")
    print(f"  Mean bias (E1 - E2 mean)     : {ov['mean_bias_eval1_minus_eval2mean']:+.3f}  "
          f"(paired t-test p={ov['paired_ttest_p']:.3f}, Wilcoxon p={ov['wilcoxon_p']:.3f})")
    print(f"  % within Evaluator-2 range   : {ov['pct_within_eval2_range']:.1%}")
    print(f"  % within ±1 SD               : {ov['pct_within_1sd']:.1%}")
    print(f"  Figures & CSVs saved to      : {RESULTS_DIR}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
